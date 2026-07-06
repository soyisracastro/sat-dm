"""Exportación de cálculos a XLSX (openpyxl) y PDF (fpdf2).

Cada calculadora construye un "documento" genérico (secciones de concepto/valor
+ tablas) y los renderers lo convierten a XLSX o PDF con los tokens de marca de
TodoConta (mismos que ``sat_descarga/procesador/exportar.py``). PTU además
genera recibos imprimibles por trabajador (sustituye al ModuloRecibosPTU.bas de
la plantilla Excel) y una hoja de pre-nómina para timbrado.

El gating premium vive en el frontend; estos renderers son cálculo puro.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .comunes import TIPOS_TERMINACION

# Tokens de diseño TodoConta (alineados con ui/src/app/globals.css).
_BRAND_PRIMARY = "0B5FFF"
_BRAND_PRIMARY_RGB = (11, 95, 255)
_BRAND_NAVY_RGB = (10, 22, 40)
_BRAND_MUTED_RGB = (107, 114, 128)
_BRAND_SUCCESS_RGB = (5, 150, 105)
_BRAND_FONT = "Calibri"

TITULOS = {
    "aguinaldo": "Aguinaldo",
    "sbc": "Salario Base de Cotización",
    "isr": "ISR de sueldos y salarios",
    "finiquito": "Finiquito",
    "liquidacion": "Liquidación",
    "carga-patronal": "Carga patronal",
    "ptu": "PTU — Reparto de utilidades",
}


def _m(v) -> str:
    """Formatea moneda MXN."""
    return f"${v:,.2f}"


def _n(v, dec: int = 2) -> str:
    return f"{v:,.{dec}f}"


# ---------------------------------------------------------------------------
# Builders de documento por calculadora
# ---------------------------------------------------------------------------


def _doc_aguinaldo(res: dict) -> dict:
    secciones = [
        {
            "titulo": "Resumen",
            "filas": [
                ("Salario diario", _m(res["salario_diario"])),
                ("Días trabajados en el año", str(res["dias_trabajados"])),
                ("Aguinaldo bruto", _m(res["aguinaldo_bruto"])),
                ("Parte exenta (30 UMA)", _m(res["parte_exenta"])),
                ("Parte gravada", _m(res["parte_gravada"])),
                ("ISR retenido", _m(res["isr_retenido"])),
                ("Tasa efectiva de ISR", f"{_n(res['tasa_efectiva_isr'])}%"),
                ("Aguinaldo neto", _m(res["aguinaldo_neto"])),
            ],
        }
    ]
    tablas = []
    comp = res.get("comparacion_metodos")
    if comp:
        recomendado = comp["metodo_recomendado"]
        tablas.append(
            {
                "titulo": "Comparación de métodos de ISR",
                "headers": ["Método", "ISR", "Tasa efectiva", "Recomendado"],
                "rows": [
                    [
                        "Art. 96 LISR (Ley)",
                        _m(comp["metodo_ley"]["isr_calculado"]),
                        f"{_n(comp['metodo_ley']['tasa_efectiva'])}%",
                        "Sí" if recomendado == "ley" else "",
                    ],
                    [
                        "Art. 174 RLISR (Reglamento)",
                        _m(comp["metodo_reglamento"]["isr_calculado"]),
                        f"{_n(comp['metodo_reglamento']['tasa_efectiva'])}%",
                        "Sí" if recomendado == "reglamento" else "",
                    ],
                ],
            }
        )
    pasos = res.get("desglose", {}).get("pasos", [])
    if pasos:
        tablas.append(
            {
                "titulo": "Desglose del cálculo",
                "headers": ["Paso", "Descripción", "Fórmula", "Resultado"],
                "rows": [
                    [
                        str(p["numero"]),
                        p["descripcion"],
                        p["formula"],
                        _m(p["resultado"]) if isinstance(p["resultado"], float) else str(p["resultado"]),
                    ]
                    for p in pasos
                ],
            }
        )
    return {"secciones": secciones, "tablas": tablas}


def _doc_sbc(res: dict) -> dict:
    d = res["desglose"]
    return {
        "secciones": [
            {
                "titulo": "Resumen",
                "filas": [
                    ("Salario diario base", _m(res["salario_diario_base"])),
                    ("Factor de integración", _n(res["factor_integracion"], 6)),
                    ("SBC diario", _m(res["sbc_diario"])),
                    ("SBC mensual", _m(res["sbc_mensual"])),
                    ("Tope legal (25 UMA)", _m(res["tope_sbc"])),
                    ("¿Excede el tope?", "Sí (se aplica el tope)" if res["excede_tope"] else "No"),
                ],
            }
        ],
        "tablas": [
            {
                "titulo": "Integración del SBC",
                "headers": ["Concepto", "Días", "Integración diaria"],
                "rows": [
                    ["Salario base", str(d["salario_base"]["dias"]), _m(d["salario_base"]["integracion_diaria"])],
                    ["Aguinaldo", str(d["aguinaldo"]["dias"]), _m(d["aguinaldo"]["integracion_diaria"])],
                    [
                        f"Prima vacacional ({_n(d['prima_vacacional']['porcentaje'], 0)}% de "
                        f"{d['prima_vacacional']['dias_vacaciones']} días)",
                        str(d["prima_vacacional"]["dias_vacaciones"]),
                        _m(d["prima_vacacional"]["integracion_diaria"]),
                    ],
                    ["Total integrado", "", _m(d["total_integrado"])],
                ],
            }
        ],
    }


def _doc_isr(res: dict) -> dict:
    d = res["desglose"]
    return {
        "secciones": [
            {
                "titulo": "Resumen",
                "filas": [
                    ("Ingreso gravado", _m(res["ingreso_bruto"])),
                    ("Periodicidad", res["periodicidad"].capitalize()),
                    ("ISR antes de subsidio", _m(res["isr_bruto"])),
                    ("Subsidio para el empleo", _m(res["subsidio_aplicado"])),
                    ("ISR final", _m(res["isr_final"])),
                    ("Tasa efectiva", f"{_n(res['tasa_efectiva'])}%"),
                    ("Ingreso neto", _m(res["ingreso_neto"])),
                ],
            },
            {
                "titulo": "Tramo de la tarifa (Art. 96 LISR)",
                "filas": [
                    ("Límite inferior", _m(d["limite_inferior"])),
                    ("Excedente del límite inferior", _m(d["excedente_limite_inferior"])),
                    ("Tasa marginal", f"{_n(d['tasa_marginal'] * 100)}%"),
                    ("Impuesto marginal", _m(d["impuesto_marginal"])),
                    ("Cuota fija", _m(d["cuota_fija"])),
                ],
            },
        ],
        "tablas": [],
    }


def _tabla_conceptos_finiquito(f: dict) -> dict:
    ag = f["aguinaldo_proporcional"]
    vac = f["vacaciones_proporcionales"]
    pv = f["prima_vacacional"]
    dev = f["salario_devengado"]
    return {
        "titulo": "Conceptos del finiquito",
        "headers": ["Concepto", "Monto", "Exento", "Gravado"],
        "rows": [
            [f"Salario devengado ({dev['dias']} días)", _m(dev["monto"]), "—", _m(dev["monto"])],
            ["Aguinaldo proporcional", _m(ag["monto"]), _m(ag["exento"]), _m(ag["gravado"])],
            [
                f"Vacaciones proporcionales ({_n(vac['dias_correspondientes'])} días)",
                _m(vac["monto"]),
                "—",
                _m(vac["monto"]),
            ],
            ["Prima vacacional", _m(pv["monto"]), _m(pv["exento"]), _m(pv["gravado"])],
        ],
    }


def _doc_finiquito(res: dict) -> dict:
    return {
        "secciones": [
            {
                "titulo": "Resumen",
                "filas": [
                    ("Antigüedad", res["antiguedad"]["texto"]),
                    ("Salario diario", _m(res["salario_diario"])),
                    ("Subtotal bruto", _m(res["subtotal_bruto"])),
                    ("Total gravado", _m(res["fiscal"]["total_gravado"])),
                    ("Total exento", _m(res["fiscal"]["total_exento"])),
                    ("ISR retenido", _m(res["total_isr"])),
                    ("Neto a pagar", _m(res["total_neto"])),
                ],
            }
        ],
        "tablas": [_tabla_conceptos_finiquito(res)],
    }


def _doc_liquidacion(res: dict) -> dict:
    tipo = TIPOS_TERMINACION.get(res["tipo_terminacion"], {})
    secciones = [
        {
            "titulo": "Resumen",
            "filas": [
                ("Tipo de terminación", tipo.get("label", res["tipo_terminacion"])),
                ("Antigüedad", res["antiguedad"]["texto"]),
                ("Salario diario", _m(res["salario_diario"])),
                ("Salario diario integrado (SDI)", _m(res["salario_diario_integrado"])),
                ("Factor de integración", _n(res["factor_integracion"], 6)),
                ("Total bruto", _m(res["total_bruto"])),
                ("Total ISR", _m(res["total_isr"])),
                ("Neto a pagar", _m(res["total_neto"])),
            ],
        }
    ]
    tablas = [_tabla_conceptos_finiquito(res["finiquito"])]

    ind = res.get("indemnizacion")
    if ind:
        fiscal_ind = res["fiscal"]["indemnizacion"]
        tablas.append(
            {
                "titulo": "Indemnización",
                "headers": ["Concepto", "Monto", "¿Aplica?"],
                "rows": [
                    [
                        "Tres meses constitucionales (Art. 48 LFT)",
                        _m(ind["tres_meses_constitucional"]["monto"]),
                        "Sí" if ind["tres_meses_constitucional"]["aplica"] else "No",
                    ],
                    [
                        "Veinte días por año (Art. 50 LFT)",
                        _m(ind["veinte_dias_por_anio"]["monto"]),
                        "Sí" if ind["veinte_dias_por_anio"]["aplica"] else "No",
                    ],
                    [
                        "Prima de antigüedad (Art. 162 LFT)",
                        _m(ind["prima_antiguedad"]["monto"]),
                        "Sí" if ind["prima_antiguedad"]["aplica"] else "No",
                    ],
                    ["Subtotal", _m(ind["subtotal"]), ""],
                    ["Exención (90 UMA por año)", _m(ind["exencion"]), ""],
                    ["Gravado", _m(ind["gravado"]), ""],
                ],
            }
        )
        secciones.append(
            {
                "titulo": "ISR de la indemnización (Art. 95 LISR)",
                "filas": [
                    ("Base gravable", _m(fiscal_ind["base_gravable"])),
                    (
                        "Método",
                        "Tasa efectiva del último sueldo"
                        if fiscal_ind["usa_tasa_efectiva"]
                        else "Tarifa directa",
                    ),
                    ("Tasa efectiva", f"{_n(fiscal_ind['tasa_efectiva'])}%"),
                    ("ISR de la indemnización", _m(fiscal_ind["isr"])),
                    ("ISR del finiquito", _m(res["fiscal"]["finiquito"]["isr"])),
                ],
            }
        )
    return {"secciones": secciones, "tablas": tablas}


def _doc_carga_patronal(res: dict) -> dict:
    return {
        "secciones": [
            {
                "titulo": "Resumen",
                "filas": [
                    ("Salario mensual", _m(res["salario_mensual"])),
                    ("SBC diario", _m(res["sbc"])),
                    ("SBC mensual", _m(res["sbc_mensual"])),
                    ("ISR del empleado", _m(res["isr_empleado"])),
                    ("Salario neto del empleado", _m(res["salario_neto"])),
                    ("Carga patronal mensual", _m(res["carga_patronal_mensual"])),
                    ("Costo total mensual", _m(res["costo_total_mensual"])),
                    ("Costo total anual", _m(res["costo_total_anual"])),
                ],
            }
        ],
        "tablas": [
            {
                "titulo": "Desglose de conceptos",
                "headers": ["Concepto", "Mensual", "Anual"],
                "rows": [
                    [c["nombre"], _m(c["monto_mensual"]), _m(c["monto_anual"])]
                    for c in res["desglose"]["conceptos"]
                ],
            }
        ],
    }


def _doc_ptu(res: dict) -> dict:
    cfg = res["config"]
    emp = res["empresa"]
    secciones = [
        {
            "titulo": "Empresa",
            "filas": [
                ("Empresa", emp["nombre"] or "—"),
                ("RFC", emp["rfc"] or "—"),
                ("Ejercicio (utilidades)", str(cfg["ejercicio"])),
                ("Año de pago", str(cfg["anio_pago"])),
                ("Utilidad fiscal", _m(emp["utilidad_fiscal"])),
                ("PTU generada (10%)", _m(emp["ptu_generada"])),
                ("PTU no cobrada de ejercicios anteriores", _m(emp["ptu_no_cobrada"])),
                ("PTU a repartir", _m(emp["ptu_a_repartir"])),
                ("Mitad repartible por días trabajados", _m(emp["bolsa_dias"])),
                ("Mitad repartible por salarios devengados", _m(emp["bolsa_salarios"])),
            ],
        },
        {
            "titulo": "Parámetros",
            "filas": [
                (
                    "Criterio de exención",
                    "UMA (criterio SAT)" if cfg["criterio_exencion"] == "UMA" else "SMG (criterio PRODECON)",
                ),
                ("Exención por trabajador (15 días)", _m(cfg["exencion_por_trabajador"])),
                ("Tipo de persona", cfg["tipo_persona"]),
                ("Fecha de pago", cfg["fecha_pago"] or "—"),
                ("Fecha límite legal", cfg["fecha_limite_pago"]),
            ],
        },
    ]
    filas_trab = []
    for t in res["trabajadores"]:
        metodo = "Art. 174" if t["comparacion"]["metodo_recomendado"] == "art174" else "Art. 96"
        filas_trab.append(
            [
                t["nombre"],
                _m(t["ptu_bruta"]),
                _m(t["monto_maximo"]),
                _m(t["ptu_real"]),
                _m(t["ptu_exenta"]),
                _m(t["ptu_gravada"]),
                _m(t["art96"]["isr_ptu"]),
                _m(t["art174"]["isr_ptu"]),
                metodo,
                _m(t["comparacion"]["isr_recomendado"]),
                _m(t["comparacion"]["ptu_neta_final"]),
            ]
        )
    tot = res["totales"]
    filas_trab.append(
        [
            "TOTALES",
            _m(tot["ptu_bruta"]),
            "",
            _m(tot["ptu_real"]),
            _m(tot["ptu_exenta"]),
            _m(tot["ptu_gravada"]),
            _m(tot["isr_art96"]),
            _m(tot["isr_art174"]),
            "",
            _m(tot["isr_recomendado"]),
            _m(tot["ptu_neta_a_pagar"]),
        ]
    )
    tablas = [
        {
            "titulo": "Reparto por trabajador",
            "headers": [
                "Trabajador", "PTU bruta", "Tope Art. 127", "PTU real", "Exenta",
                "Gravada", "ISR Art. 96", "ISR Art. 174", "Método", "ISR retenido", "PTU neta",
            ],
            "rows": filas_trab,
        },
        {
            "titulo": "Pre-nómina (borrador para timbrado CFDI)",
            "hoja": "Pre_Nómina",
            "headers": [
                "Trabajador", "RFC", "CURP", "NSS", "Régimen", "Tipo nómina",
                "Clave perc.", "PTU gravada", "PTU exenta", "Clave ded.", "ISR retenido", "Neto a pagar",
            ],
            "rows": [
                [
                    t["nombre"], t["rfc"] or "—", t["curp"] or "—", t["nss"] or "—",
                    t["prenomina"]["regimen"], t["prenomina"]["tipo_nomina"],
                    t["prenomina"]["clave_percepcion"], _m(t["prenomina"]["ptu_gravada"]),
                    _m(t["prenomina"]["ptu_exenta"]), t["prenomina"]["clave_deduccion"],
                    _m(t["prenomina"]["isr_retenido"]), _m(t["prenomina"]["neto_a_pagar"]),
                ]
                for t in res["trabajadores"]
            ],
        },
    ]
    advertencias = list(res.get("advertencias", []))
    for t in res["trabajadores"]:
        advertencias += [f"{t['nombre']}: {a}" for a in t.get("advertencias", [])]
    return {"secciones": secciones, "tablas": tablas, "advertencias": advertencias}


_BUILDERS = {
    "aguinaldo": _doc_aguinaldo,
    "sbc": _doc_sbc,
    "isr": _doc_isr,
    "finiquito": _doc_finiquito,
    "liquidacion": _doc_liquidacion,
    "carga-patronal": _doc_carga_patronal,
    "ptu": _doc_ptu,
}


def construir_documento(calculadora: str, resultado: dict, anio: int) -> dict:
    if calculadora not in _BUILDERS:
        raise ValueError(f"Exportación no soportada para {calculadora!r}.")
    doc = _BUILDERS[calculadora](resultado)
    doc["titulo"] = TITULOS[calculadora]
    doc["subtitulo"] = f"TodoConta · Ejercicio {anio}"
    doc.setdefault("advertencias", [])
    return doc


# ---------------------------------------------------------------------------
# Render XLSX
# ---------------------------------------------------------------------------

_FILL_HEADER = PatternFill(start_color=_BRAND_PRIMARY, end_color=_BRAND_PRIMARY, fill_type="solid")
_FONT_HEADER = Font(name=_BRAND_FONT, bold=True, color="FFFFFF", size=11)
_FONT_TITULO = Font(name=_BRAND_FONT, bold=True, size=15, color="0A1628")
_FONT_SUB = Font(name=_BRAND_FONT, size=10, color="6B7280")
_FONT_SECCION = Font(name=_BRAND_FONT, bold=True, size=12, color=_BRAND_PRIMARY)
_FONT_LABEL = Font(name=_BRAND_FONT, bold=True, size=11)
_FONT_NORMAL = Font(name=_BRAND_FONT, size=11)


def _ancho_columnas(ws, rows: list[list[str]], minimo: int = 12, maximo: int = 48) -> None:
    from openpyxl.utils import get_column_letter

    anchos: dict[int, int] = {}
    for row in rows:
        for i, celda in enumerate(row, 1):
            anchos[i] = max(anchos.get(i, minimo), min(len(str(celda)) + 2, maximo))
    for i, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _escribir_tabla(ws, fila: int, tabla: dict) -> int:
    ws.cell(row=fila, column=1, value=tabla["titulo"]).font = _FONT_SECCION
    fila += 1
    for ci, h in enumerate(tabla["headers"], 1):
        c = ws.cell(row=fila, column=ci, value=h)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fila += 1
    for row in tabla["rows"]:
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=fila, column=ci, value=v)
            c.font = _FONT_NORMAL
            if ci > 1:
                c.alignment = Alignment(horizontal="right")
        fila += 1
    return fila + 1


def a_xlsx(doc: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    ws.cell(row=1, column=1, value=doc["titulo"]).font = _FONT_TITULO
    ws.cell(row=2, column=1, value=doc["subtitulo"]).font = _FONT_SUB
    fila = 4
    for seccion in doc["secciones"]:
        ws.cell(row=fila, column=1, value=seccion["titulo"]).font = _FONT_SECCION
        fila += 1
        for etiqueta, valor in seccion["filas"]:
            ws.cell(row=fila, column=1, value=etiqueta).font = _FONT_LABEL
            c = ws.cell(row=fila, column=2, value=valor)
            c.font = _FONT_NORMAL
            c.alignment = Alignment(horizontal="right")
            fila += 1
        fila += 1
    if doc["advertencias"]:
        ws.cell(row=fila, column=1, value="Advertencias").font = _FONT_SECCION
        fila += 1
        for a in doc["advertencias"]:
            ws.cell(row=fila, column=1, value=f"• {a}").font = _FONT_SUB
            fila += 1
    _ancho_columnas(ws, [[f, v] for s in doc["secciones"] for f, v in s["filas"]], minimo=24)

    principales = [t for t in doc["tablas"] if not t.get("hoja")]
    if principales:
        ws2 = wb.create_sheet("Desglose")
        fila2 = 1
        todas_las_filas: list[list[str]] = []
        for tabla in principales:
            fila2 = _escribir_tabla(ws2, fila2, tabla)
            todas_las_filas += [tabla["headers"]] + tabla["rows"]
        _ancho_columnas(ws2, todas_las_filas)

    for tabla in doc["tablas"]:
        if tabla.get("hoja"):
            ws3 = wb.create_sheet(tabla["hoja"])
            _escribir_tabla(ws3, 1, tabla)
            _ancho_columnas(ws3, [tabla["headers"]] + tabla["rows"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Render PDF (fpdf2, fuentes core Helvetica → sanitizar a latin-1)
# ---------------------------------------------------------------------------


def _latin1(texto: str) -> str:
    return str(texto).encode("latin-1", errors="replace").decode("latin-1")


def _pdf_base(titulo: str, subtitulo: str):
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="Letter")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*_BRAND_PRIMARY_RGB)
    pdf.cell(0, 6, "TodoConta", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(*_BRAND_NAVY_RGB)
    pdf.cell(0, 9, _latin1(titulo), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*_BRAND_MUTED_RGB)
    pdf.cell(0, 6, _latin1(subtitulo), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    return pdf


def _pdf_seccion(pdf, seccion: dict) -> None:
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*_BRAND_PRIMARY_RGB)
    pdf.cell(0, 8, _latin1(seccion["titulo"]), new_x="LMARGIN", new_y="NEXT")
    ancho = pdf.w - pdf.l_margin - pdf.r_margin
    for etiqueta, valor in seccion["filas"]:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*_BRAND_NAVY_RGB)
        pdf.cell(ancho * 0.55, 6, _latin1(etiqueta))
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(ancho * 0.45, 6, _latin1(valor), align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)


def _pdf_tabla(pdf, tabla: dict) -> None:
    n = len(tabla["headers"])
    if n > 6 or pdf.get_y() > pdf.h - 60:
        pdf.add_page(orientation="L" if n > 6 else pdf.cur_orientation)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*_BRAND_PRIMARY_RGB)
    pdf.cell(0, 8, _latin1(tabla["titulo"]), new_x="LMARGIN", new_y="NEXT")

    ancho_total = pdf.w - pdf.l_margin - pdf.r_margin
    # Primera columna más ancha (nombres/conceptos).
    ancho_primera = ancho_total * (0.30 if n > 3 else 0.4)
    ancho_resto = (ancho_total - ancho_primera) / (n - 1) if n > 1 else ancho_total
    anchos = [ancho_primera] + [ancho_resto] * (n - 1)

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(255, 255, 255)
    pdf.set_fill_color(*_BRAND_PRIMARY_RGB)
    for w, h in zip(anchos, tabla["headers"]):
        pdf.cell(w, 7, _latin1(h), fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_BRAND_NAVY_RGB)
    relleno = False
    pdf.set_fill_color(247, 249, 252)
    for row in tabla["rows"]:
        for i, (w, v) in enumerate(zip(anchos, row)):
            texto = _latin1(v)
            # Truncado defensivo para que la celda no desborde.
            max_chars = max(6, int(w / 1.7))
            if len(texto) > max_chars:
                texto = texto[: max_chars - 1] + "…".encode("latin-1", "replace").decode("latin-1")
            pdf.cell(w, 6, texto, fill=relleno, align="L" if i == 0 else "R")
        pdf.ln()
        relleno = not relleno
    pdf.ln(3)


def a_pdf(doc: dict) -> bytes:
    pdf = _pdf_base(doc["titulo"], doc["subtitulo"])
    for seccion in doc["secciones"]:
        _pdf_seccion(pdf, seccion)
    for tabla in doc["tablas"]:
        _pdf_tabla(pdf, tabla)
    if doc["advertencias"]:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(180, 83, 9)
        pdf.cell(0, 8, "Advertencias", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        for a in doc["advertencias"]:
            pdf.multi_cell(0, 5, _latin1(f"- {a}"), new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Recibos de PTU (uno por trabajador, listos para firma)
# ---------------------------------------------------------------------------


def recibos_ptu_pdf(resultado: dict) -> bytes:
    """PDF multi-página: un recibo de PTU por trabajador (desglose + firma)."""
    from fpdf import FPDF

    cfg = resultado["config"]
    emp = resultado["empresa"]

    pdf = FPDF(orientation="P", unit="mm", format="Letter")
    pdf.set_auto_page_break(auto=False)

    for t in resultado["trabajadores"]:
        pdf.add_page()

        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*_BRAND_PRIMARY_RGB)
        pdf.cell(0, 9, _latin1(emp["nombre"] or "Recibo de PTU"), align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*_BRAND_MUTED_RGB)
        pdf.cell(0, 6, _latin1(emp["rfc"] or ""), align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*_BRAND_PRIMARY_RGB)
        pdf.cell(0, 8, _latin1(f"RECIBO DE PTU — Ejercicio {cfg['ejercicio']}"), align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

        datos = [
            ("Trabajador", t["nombre"]),
            ("RFC", t["rfc"] or "—"),
            ("CURP", t["curp"] or "—"),
            ("Fecha de pago", cfg["fecha_pago"] or "—"),
        ]
        ancho = pdf.w - pdf.l_margin - pdf.r_margin
        for etiqueta, valor in datos:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*_BRAND_NAVY_RGB)
            pdf.cell(ancho * 0.3, 6, _latin1(f"{etiqueta}:"))
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(ancho * 0.7, 6, _latin1(valor), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

        metodo = "Art. 174 RLISR" if t["comparacion"]["metodo_recomendado"] == "art174" else "Art. 96 LISR"
        desglose = [
            ("PTU bruta", _m(t["ptu_bruta"])),
            ("Tope aplicado (Art. 127 fr. VIII LFT)", _m(t["monto_maximo"])),
            ("PTU real", _m(t["ptu_real"])),
            ("PTU exenta", _m(t["ptu_exenta"])),
            ("PTU gravada", _m(t["ptu_gravada"])),
            ("ISR retenido", _m(t["comparacion"]["isr_recomendado"])),
            ("Método de ISR", metodo),
        ]
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(255, 255, 255)
        pdf.set_fill_color(*_BRAND_PRIMARY_RGB)
        pdf.cell(ancho * 0.6, 7, "Concepto", fill=True)
        pdf.cell(ancho * 0.4, 7, "Importe", fill=True, align="R", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*_BRAND_NAVY_RGB)
        relleno = False
        pdf.set_fill_color(247, 249, 252)
        for etiqueta, valor in desglose:
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(ancho * 0.6, 7, _latin1(etiqueta), fill=relleno)
            pdf.cell(ancho * 0.4, 7, _latin1(valor), fill=relleno, align="R", new_x="LMARGIN", new_y="NEXT")
            relleno = not relleno

        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*_BRAND_SUCCESS_RGB)
        pdf.cell(ancho * 0.6, 9, "PTU NETA A RECIBIR")
        pdf.cell(ancho * 0.4, 9, _latin1(_m(t["comparacion"]["ptu_neta_final"])), align="R", new_x="LMARGIN", new_y="NEXT")

        pdf.ln(14)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*_BRAND_MUTED_RGB)
        pdf.cell(0, 6, _latin1("Recibí de conformidad la cantidad arriba señalada."), align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(12)
        pdf.cell(0, 6, "_" * 60, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, _latin1("Nombre y firma del trabajador"), align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(
            0, 5,
            _latin1("De conformidad con los artículos 117 al 131 de la Ley Federal del Trabajo."),
            align="C", new_x="LMARGIN", new_y="NEXT",
        )

    return bytes(pdf.output())
