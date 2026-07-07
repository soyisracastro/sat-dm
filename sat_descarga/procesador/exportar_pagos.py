"""
Exportación XLSX multi-sheet del procesador de Pagos.

Sheets:
1. Resumen        — métricas ejecutivas (siempre).
2. Facturas PPD   — UUID, monto, status, # pagos (si hay PPDs).
3. Detalle pagos  — relación 1:N PPD ↔ complementos (si hay relaciones).
4. Análisis fechas — extemporáneos con días de retraso (condicional).
5. Pagos huérfanos — complementos sin PPD cargada (condicional).
6. Incidencias PUE — PUE+complemento (condicional, estilo alerta roja).

Reusa los tokens de diseño TodoConta (`#0B5FFF` + Calibri) definidos en
`exportar.py:_BRAND_*` para consistencia con el XLSX del procesador CFDI.
"""

from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import reportes_pagos as rep
from .db import ProcesadorDB
from .exportar import (
    _BRAND_FONT_NAME,
    _BRAND_FONT_SIZE,
    _BRAND_PRIMARY,
    _BRAND_PRIMARY_FOREGROUND,
    _header_row,
)

# Tono para sheets de alertas (huérfanos, incidencias PUE) — rojo destructive
# del design system pero más opaco para legibilidad sobre blanco.
_ALERT_FILL = "B91C1C"


_STATUS_LABEL = {
    "sin_complemento": "Sin complemento",
    "pago_parcial": "Pago parcial",
    "pagado_completo": "Pagado completo",
    "sobrante": "Sobrante",
}


def to_xlsx(db: ProcesadorDB, filtros: Optional[dict] = None) -> bytes:
    """Genera el XLSX completo del procesador de Pagos."""
    wb = Workbook(write_only=True)

    header_font = Font(
        bold=True,
        color=_BRAND_PRIMARY_FOREGROUND,
        name=_BRAND_FONT_NAME,
        size=_BRAND_FONT_SIZE,
    )
    header_fill = PatternFill("solid", fgColor=_BRAND_PRIMARY)
    alert_font = Font(
        bold=True,
        color=_BRAND_PRIMARY_FOREGROUND,
        name=_BRAND_FONT_NAME,
        size=_BRAND_FONT_SIZE,
    )
    alert_fill = PatternFill("solid", fgColor=_ALERT_FILL)

    stats = rep.stats_pagos(db, filtros)
    facturas = rep.facturas_ppd(db, filtros, page=1, page_size=100_000)["items"]
    extemp = rep.analisis_fechas(db, filtros)
    huerfanos = rep.pagos_huerfanos(db, filtros)
    pue = rep.incidencias_pue(db, filtros)

    # ---------- 1. Resumen ----------
    ws = wb.create_sheet("Resumen")
    ws.freeze_panes = "A2"
    ws.append(_header_row(ws, ["Métrica", "Valor"], header_font, header_fill))
    for k, label in [
        ("total_ingresos_ppd", "Total facturas PPD"),
        ("porcentaje_conciliados", "% conciliadas"),
        ("sin_complemento", "Sin complemento"),
        ("pagos_parciales", "Pagos parciales"),
        ("pagos_completos", "Pagos completos"),
        ("sobrantes", "Sobrantes"),
        ("monto_total_sin_pagar", "Monto total sin pagar"),
        ("total_pagos", "Total complementos (tipo P)"),
        ("pagos_huerfanos", "Pagos huérfanos"),
        ("incidencias_pue", "Incidencias PUE+complemento"),
        ("complementos_extemporaneos", "Complementos extemporáneos"),
        ("monto_complementos_extemporaneos", "Monto extemporáneos"),
    ]:
        ws.append([label, stats.get(k)])

    # ---------- 2. Facturas PPD ----------
    if facturas:
        ws = wb.create_sheet("Facturas PPD")
        ws.freeze_panes = "A2"
        cols = [
            "UUID", "Fecha", "Serie", "Folio",
            "RFC Emisor", "Nombre Emisor",
            "RFC Receptor", "Nombre Receptor",
            "Total", "Pagado", "Saldo pendiente", "# pagos",
            "Moneda", "Status", "Estado SAT", "Warnings",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for f in facturas:
            ws.append([
                f["uuid"], f["fecha"], f["serie"], f["folio"],
                f["emisor_rfc"], f["emisor_nombre"],
                f["receptor_rfc"], f["receptor_nombre"],
                f["total"], f["total_pagado"], f["saldo_pendiente"], f["num_pagos"],
                f["moneda"], _STATUS_LABEL.get(f["status"], f["status"]),
                f.get("estado_sat") or "Sin validar",
                " | ".join(f["warnings"] or []),
            ])

    # ---------- 3. Detalle de pagos ----------
    if facturas:
        ws = wb.create_sheet("Detalle de pagos")
        ws.freeze_panes = "A2"
        cols = [
            "UUID Factura", "Folio Factura", "Total Factura",
            "UUID Complemento", "Fecha pago", "Monto pagado",
            "Forma pago", "Parcialidad", "Saldo anterior", "Saldo insoluto",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for f in facturas:
            if f["num_pagos"] == 0:
                continue
            for p in rep.detalle_pagos_de_ppd(db, f["uuid"], (filtros or {}).get("mi_rfc")):
                ws.append([
                    f["uuid"], f["folio"], f["total"],
                    p["cfdi_pago_uuid"], p["cfdi_pago_fecha_pago"],
                    p["docto_imp_pagado"], p["cfdi_pago_forma"],
                    p["docto_num_parcialidad"],
                    p["docto_imp_saldo_ant"], p["docto_imp_saldo_insoluto"],
                ])

    # ---------- 4. Análisis de fechas (extemporáneos) ----------
    if extemp:
        ws = wb.create_sheet("Análisis fechas")
        ws.freeze_panes = "A2"
        cols = [
            "UUID Complemento", "Fecha emisión", "Fecha pago",
            "Fecha límite", "Días de retraso",
            "Monto", "RFC Emisor", "Nombre Emisor",
            "UUID Factura", "Folio Factura",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for e in extemp:
            ws.append([
                e["cfdi_pago_uuid"],
                e["fecha_emision_complemento"],
                e["cfdi_pago_fecha_pago"],
                e["limite"],
                e["dias_retraso"],
                e["monto_complemento"],
                e["emisor_rfc"], e["emisor_nombre"],
                e["factura_uuid"], e["factura_folio"],
            ])

    # ---------- 5. Pagos huérfanos ----------
    if huerfanos:
        ws = wb.create_sheet("Pagos huérfanos")
        ws.freeze_panes = "A2"
        cols = [
            "UUID Pago", "Fecha emisión",
            "RFC Emisor", "Nombre Emisor",
            "Monto", "Documentos referenciados",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for h in huerfanos:
            ws.append([
                h["cfdi_pago_uuid"], h["fecha_emision"],
                h["emisor_rfc"], h["emisor_nombre"],
                h["monto"], h.get("documentos_referenciados") or "",
            ])

    # ---------- 6. Incidencias PUE ----------
    if pue:
        ws = wb.create_sheet("Incidencias PUE")
        ws.freeze_panes = "A2"
        cols = [
            "UUID Factura PUE", "Fecha Factura",
            "RFC Emisor", "Nombre Emisor",
            "Total Factura", "Método Factura",
            "UUID Complemento", "Fecha pago", "Monto pagado",
            "Descripción del riesgo",
        ]
        ws.append(_header_row(ws, cols, alert_font, alert_fill))
        for i in pue:
            ws.append([
                i.get("factura_uuid"), i.get("factura_fecha"),
                i.get("emisor_rfc"), i.get("emisor_nombre"),
                i.get("factura_total"), i.get("factura_metodo_pago"),
                i.get("complemento_uuid"), i.get("cfdi_pago_fecha_pago"),
                i.get("monto_pagado"),
                i.get("descripcion_riesgo"),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
