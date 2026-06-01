"""
Exportación del procesador a Excel (XLSX) y CSV.

Excel: dos sheets — `CFDIs` (una fila por comprobante) y `Conceptos` (una fila
por concepto, ligada a UUID padre). Usa openpyxl en modo `write_only` para
streaming con bajo uso de memoria.

CSV: una fila por CFDI, con BOM UTF-8 para compatibilidad con Excel para
Windows. Opcionalmente expande conceptos.
"""

from __future__ import annotations

import csv
import io
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .db import CfdiFiltros, ProcesadorDB


# Tokens de diseño TodoConta (alineados con ui/src/app/globals.css).
# Se aplica al XLSX para que los documentos exportados sean visualmente
# consistentes con la app web.
#
# Tipografía: la app usa Inter, pero como Excel no la trae por default
# usamos Calibri (sans-serif moderna, universal en Windows y macOS Office
# desde 2007) — es el equivalente más cercano a Inter disponible en todos
# los sistemas sin requerir instalación.
#
# Colores: hex literal de la paleta TodoConta. Excel los renderiza tal
# cual (es el mismo formato que usa CSS).
_BRAND_PRIMARY = "0B5FFF"            # --primary (azul TodoConta)
_BRAND_PRIMARY_FOREGROUND = "FFFFFF"  # --primary-foreground (blanco)
_BRAND_FONT_NAME = "Calibri"          # universal; fallback más cercano a Inter
_BRAND_FONT_SIZE = 11


# Columnas del sheet "CFDIs" (y del CSV plano)
_COLS_CFDI = [
    ("uuid", "UUID"),
    ("tipo", "Tipo"),
    ("fecha", "Fecha"),
    ("serie", "Serie"),
    ("folio", "Folio"),
    ("emisor_rfc", "RFC Emisor"),
    ("emisor_nombre", "Nombre Emisor"),
    ("receptor_rfc", "RFC Receptor"),
    ("receptor_nombre", "Nombre Receptor"),
    ("sub_total", "Subtotal"),
    ("descuento", "Descuento"),
    ("iva_trasladado", "IVA Trasladado"),
    ("ieps_trasladado", "IEPS Trasladado"),
    ("iva_retenido", "IVA Retenido"),
    ("isr_retenido", "ISR Retenido"),
    ("total", "Total"),
    ("moneda", "Moneda"),
    ("tipo_cambio", "Tipo de Cambio"),
    ("forma_pago", "Forma de Pago"),
    ("metodo_pago", "Método de Pago"),
    ("receptor_uso_cfdi", "Uso CFDI"),
    ("lugar_expedicion", "Lugar Expedición"),
    ("estado_sat", "Estado SAT"),
]


_COLS_CONCEPTO = [
    ("cfdi_uuid", "UUID CFDI"),
    ("clave_prod_serv", "Clave Prod/Serv"),
    ("descripcion", "Descripción"),
    ("cantidad", "Cantidad"),
    ("clave_unidad", "Clave Unidad"),
    ("unidad", "Unidad"),
    ("valor_unitario", "Valor Unitario"),
    ("importe", "Importe"),
    ("descuento", "Descuento"),
]


def to_xlsx(
    db: ProcesadorDB,
    filtros: Optional[CfdiFiltros] = None,
) -> bytes:
    """
    Genera un XLSX con dos sheets (`CFDIs` y `Conceptos`) y devuelve los bytes.
    Modo write_only para bajo uso de memoria con +1000 CFDIs.
    """
    wb = Workbook(write_only=True)
    header_font = Font(
        bold=True,
        color=_BRAND_PRIMARY_FOREGROUND,
        name=_BRAND_FONT_NAME,
        size=_BRAND_FONT_SIZE,
    )
    header_fill = PatternFill("solid", fgColor=_BRAND_PRIMARY)

    # Sheet CFDIs
    ws_cfdis = wb.create_sheet("CFDIs")
    ws_cfdis.freeze_panes = "A2"
    headers = [label for _, label in _COLS_CFDI]
    ws_cfdis.append(_header_row(ws_cfdis, headers, font=header_font, fill=header_fill))

    uuids_cargados: list[str] = []
    for cfdi in db.iter_all(filtros):
        ws_cfdis.append([cfdi.get(key) for key, _ in _COLS_CFDI])
        uuids_cargados.append(cfdi["uuid"])

    # Sheet Conceptos — cargado a partir de los UUIDs ya filtrados
    ws_conceptos = wb.create_sheet("Conceptos")
    ws_conceptos.freeze_panes = "A2"
    headers_c = [label for _, label in _COLS_CONCEPTO]
    ws_conceptos.append(_header_row(ws_conceptos, headers_c, font=header_font, fill=header_fill))

    if uuids_cargados:
        with db.cursor() as cur:
            # Chunks por límite SQLite de placeholders (999 por default)
            for chunk in _chunks(uuids_cargados, 500):
                placeholders = ",".join("?" for _ in chunk)
                cur.execute(
                    f"""
                    SELECT cfdi_uuid, clave_prod_serv, descripcion, cantidad,
                           clave_unidad, unidad, valor_unitario, importe, descuento
                    FROM conceptos
                    WHERE cfdi_uuid IN ({placeholders})
                    """,
                    chunk,
                )
                for row in cur.fetchall():
                    ws_conceptos.append([row[k] for k, _ in _COLS_CONCEPTO])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def to_csv(
    db: ProcesadorDB,
    filtros: Optional[CfdiFiltros] = None,
    expandir_conceptos: bool = False,
) -> bytes:
    """
    Genera CSV con BOM UTF-8 (Excel-compat). Si `expandir_conceptos` es True,
    se emite una fila por concepto incluyendo las columnas del CFDI padre.
    """
    buf = io.StringIO()
    buf.write("﻿")  # BOM
    writer = csv.writer(buf)

    if not expandir_conceptos:
        writer.writerow([label for _, label in _COLS_CFDI])
        for cfdi in db.iter_all(filtros):
            writer.writerow([cfdi.get(key, "") for key, _ in _COLS_CFDI])
    else:
        headers = (
            [label for _, label in _COLS_CFDI]
            + [label for k, label in _COLS_CONCEPTO if k != "cfdi_uuid"]
        )
        writer.writerow(headers)
        for cfdi in db.iter_all(filtros):
            base = [cfdi.get(key, "") for key, _ in _COLS_CFDI]
            conceptos = db.conceptos_de(cfdi["uuid"])
            if not conceptos:
                writer.writerow(base + [""] * (len(_COLS_CONCEPTO) - 1))
            else:
                for c in conceptos:
                    writer.writerow(
                        base + [c.get(key, "") for key, _ in _COLS_CONCEPTO if key != "cfdi_uuid"]
                    )

    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _chunks(seq: list, size: int) -> Iterable[list]:
    for i in range(0, len(seq), size):
        yield seq[i : i + size]


def _header_row(ws, values: list, font: Font, fill: PatternFill) -> list:
    """
    Construye la fila de encabezado con la paleta TodoConta (fondo `--primary`
    azul `#0B5FFF`, texto blanco, fuente Inter bold). El sheet en modo
    write_only requiere `WriteOnlyCell` con referencia explícita al ws.
    """
    from openpyxl.cell import WriteOnlyCell

    align = Alignment(horizontal="left", vertical="center")
    cells = []
    for v in values:
        c = WriteOnlyCell(ws, value=v)
        c.font = font
        c.fill = fill
        c.alignment = align
        cells.append(c)
    return cells
