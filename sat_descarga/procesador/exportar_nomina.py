"""
Exportación XLSX multi-sheet del procesador de Nómina.

Sheets (en orden):
1. Disclaimer                — leyenda fiscal + fecha + versión (siempre, primer sheet).
2. Resumen                   — KPIs aplanados (siempre).
3. Recibos                   — 1 fila por CFDI tipo N con datos del trabajador y totales.
4. Percepciones              — 1 fila por concepto Percepcion (gravado/exento desglosado).
5. Deducciones               — 1 fila por concepto Deduccion.
6. Otros Pagos               — 1 fila por OtroPago (con subsidio causado para tipo 002).
7. Deductibilidad — Global   — métricas ejecutivas de ISR teórico vs retenido.
8. Deductibilidad — Por empleado — desglose por RFC con advertencias de periodo.
9. IMSS                      — registros con SBC, SDI, aportaciones, observaciones.
10. IMSS — Alertas           — solo si hay alertas (rojo destructive).
11. Periodo vs Periodo       — solo si hay datos suficientes.

Reusa los tokens TodoConta `#0B5FFF + Calibri` y `_ALERT_FILL` (rojo destructive).
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from . import reportes_nomina as rep
from .catalogos import (
    PERIODICIDAD_PAGO,
    TIPO_DEDUCCION,
    TIPO_NOMINA,
    TIPO_OTRO_PAGO,
    TIPO_PERCEPCION,
)
from .db import ProcesadorDB
from .exportar import (
    _BRAND_FONT_NAME,
    _BRAND_FONT_SIZE,
    _BRAND_PRIMARY,
    _BRAND_PRIMARY_FOREGROUND,
    _header_row,
)
from .exportar_pagos import _ALERT_FILL


DISCLAIMER_TEXTO = (
    "Los cálculos de ISR teórico y aportaciones IMSS son orientativos, "
    "basados en tarifas oficiales del año fiscal detectado en los recibos. "
    "No reemplazan el cálculo de tu sistema de nómina ni la opinión de un "
    "contador. Confirma siempre con un profesional antes de presentar "
    "declaraciones o pagos al SAT/IMSS."
)


def _label_tipo(catalogo: dict, code: str) -> str:
    if not code:
        return ""
    desc = catalogo.get(code)
    return f"{code} — {desc}" if desc else code


def to_xlsx(db: ProcesadorDB, filtros: Optional[dict] = None) -> bytes:
    """Genera el XLSX completo del procesador de Nómina."""
    wb = Workbook(write_only=True)

    header_font = Font(
        bold=True,
        color=_BRAND_PRIMARY_FOREGROUND,
        name=_BRAND_FONT_NAME,
        size=_BRAND_FONT_SIZE,
    )
    header_fill = PatternFill("solid", fgColor=_BRAND_PRIMARY)
    alert_font = Font(
        bold=True,
        color=_BRAND_PRIMARY_FOREGROUND,
        name=_BRAND_FONT_NAME,
        size=_BRAND_FONT_SIZE,
    )
    alert_fill = PatternFill("solid", fgColor=_ALERT_FILL)

    # Cargar todos los datos.
    stats = rep.stats_nomina(db, filtros)
    recibos = rep.listar_recibos(db, filtros, page=1, page_size=100_000)["items"]
    deduc = rep.reporte_deducibilidad(db, filtros)
    imss = rep.reporte_imss(db, filtros)
    pvsp = rep.reporte_periodo_vs_periodo(db, filtros)

    # Conceptos planos (1 fila por concepto): se reutiliza el helper de
    # reportes para obtener el shape aplanado.
    all_records = rep._cargar_records(db, filtros)

    # ---------- 1. Disclaimer (PRIMER sheet) ----------
    ws = wb.create_sheet("Disclaimer")
    ws.append(_header_row(ws, ["Aviso fiscal"], header_font, header_fill))
    ws.append([DISCLAIMER_TEXTO])
    ws.append([])
    ws.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Procesador", "sat-descarga-masiva — Nómina"])

    # ---------- 2. Resumen ----------
    ws = wb.create_sheet("Resumen")
    ws.freeze_panes = "A2"
    ws.append(_header_row(ws, ["Métrica", "Valor"], header_font, header_fill))
    for k, label in [
        ("total_recibos", "Total recibos"),
        ("total_empleados", "Empleados únicos"),
        ("total_conceptos", "Total conceptos"),
        ("nominas_ordinarias", "Nóminas ordinarias"),
        ("nominas_extraordinarias", "Nóminas extraordinarias"),
        ("total_percepciones", "Total percepciones"),
        ("total_deducciones", "Total deducciones"),
        ("total_otros_pagos", "Total otros pagos"),
        ("neto_a_pagar", "Neto a pagar"),
        ("conceptos_con_errores", "Recibos con warnings"),
    ]:
        ws.append([label, stats.get(k)])

    # ---------- 3. Recibos ----------
    if recibos:
        ws = wb.create_sheet("Recibos")
        ws.freeze_panes = "A2"
        cols = [
            "UUID", "Fecha pago", "Periodo (inicio)", "Periodo (fin)",
            "RFC Empleado", "Nombre", "NSS", "CURP", "Núm. empleado", "Puesto",
            "Periodicidad", "Tipo nómina", "Días pagados",
            "SBC", "SDI",
            "Total percepciones", "Total deducciones", "Total otros pagos", "Neto",
            "Estado SAT",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for r in recibos:
            ws.append([
                r["cfdi_uuid"],
                r.get("fecha_pago"),
                r.get("fecha_inicial_pago"),
                r.get("fecha_final_pago"),
                r.get("receptor_rfc"),
                r.get("receptor_nombre"),
                r.get("nss") or "",
                r.get("curp") or "",
                r.get("num_empleado") or "",
                r.get("puesto") or "",
                PERIODICIDAD_PAGO.get(r.get("periodicidad_pago") or "", r.get("periodicidad_pago") or ""),
                TIPO_NOMINA.get(r.get("tipo_nomina") or "", r.get("tipo_nomina") or ""),
                r.get("num_dias_pagados"),
                r.get("salario_base_cot_apor"),
                r.get("salario_diario_integrado"),
                r.get("total_percepciones"),
                r.get("total_deducciones"),
                r.get("total_otros_pagos"),
                r.get("neto"),
                r.get("estado_sat") or "Sin validar",
            ])

    # Conceptos divididos por clase.
    percepciones = [r for r in all_records if r["clase"] == "Percepcion"]
    deducciones = [r for r in all_records if r["clase"] == "Deduccion"]
    otros = [r for r in all_records if r["clase"] == "OtroPago"]

    # ---------- 4. Percepciones ----------
    if percepciones:
        ws = wb.create_sheet("Percepciones")
        ws.freeze_panes = "A2"
        cols = [
            "UUID", "RFC Empleado", "Nombre", "Tipo", "Concepto",
            "Importe gravado", "Importe exento", "Total",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for p in percepciones:
            total = (p["importe_gravado"] or 0.0) + (p["importe_exento"] or 0.0)
            ws.append([
                p["cfdi_uuid"], p["receptor_rfc"], p["receptor_nombre"],
                _label_tipo(TIPO_PERCEPCION, p.get("tipo_concepto") or ""),
                p.get("concepto"),
                p["importe_gravado"], p["importe_exento"], total,
            ])

    # ---------- 5. Deducciones ----------
    if deducciones:
        ws = wb.create_sheet("Deducciones")
        ws.freeze_panes = "A2"
        cols = ["UUID", "RFC Empleado", "Nombre", "Tipo", "Concepto", "Importe"]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for d in deducciones:
            ws.append([
                d["cfdi_uuid"], d["receptor_rfc"], d["receptor_nombre"],
                _label_tipo(TIPO_DEDUCCION, d.get("tipo_concepto") or ""),
                d.get("concepto"),
                d["importe"],
            ])

    # ---------- 6. Otros Pagos ----------
    if otros:
        ws = wb.create_sheet("Otros Pagos")
        ws.freeze_panes = "A2"
        cols = [
            "UUID", "RFC Empleado", "Nombre", "Tipo", "Concepto",
            "Importe", "Subsidio causado",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for o in otros:
            ws.append([
                o["cfdi_uuid"], o["receptor_rfc"], o["receptor_nombre"],
                _label_tipo(TIPO_OTRO_PAGO, o.get("tipo_concepto") or ""),
                o.get("concepto"),
                o["importe"], o["subsidio_causado"],
            ])

    # ---------- 7. Deductibilidad — Global ----------
    ws = wb.create_sheet("Deductibilidad — Global")
    ws.freeze_panes = "A2"
    ws.append(_header_row(ws, ["Métrica", "Valor"], header_font, header_fill))
    analisis = deduc["isr_analisis"]
    for label, value in [
        ("Periodo inicio", deduc["periodo_inicio"]),
        ("Periodo fin", deduc["periodo_fin"]),
        ("Tarifa aplicada", analisis["tarifa_label"]),
        ("Año detectado", analisis["year_detected"]),
        ("Total percepciones", deduc["total_percepciones"]),
        ("Percepciones gravadas", deduc["percepciones_gravadas"]),
        ("Percepciones exentas", deduc["percepciones_exentas"]),
        ("Total deducciones", deduc["total_deducciones"]),
        ("Seguro social retenido", deduc["seguro_social"]),
        ("ISR retenido", deduc["isr_retenido"]),
        ("ISR bruto (calculado)", analisis["isr_bruto"]),
        ("Subsidio al empleo aplicado", analisis["subsidio_aplicado"]),
        ("ISR teórico (calculado)", analisis["isr_teorico"]),
        ("ISR diferencia (retenido − teórico)", analisis["isr_diferencia"]),
        ("Aporta. retiro/cesantía", deduc["aportaciones_retiro_cesantia"]),
        ("Otros deducciones", deduc["otros_deducciones"]),
        ("Salario neto", deduc["salario_neto"]),
        ("Empleados analizados", deduc["empleados_analizados"]),
        ("Cumplimiento retenciones", deduc["detalle_analisis"]["cumplimiento_retenciones"]),
        ("Adecuación fiscal", deduc["detalle_analisis"]["adecuacion_fiscal"]),
    ]:
        ws.append([label, value])

    if deduc.get("recomendaciones"):
        ws.append([])
        ws.append(["Recomendaciones"])
        for rec in deduc["recomendaciones"]:
            ws.append([rec])

    if deduc.get("advertencias_periodo"):
        ws.append([])
        ws.append(["Advertencias de periodo"])
        for w in deduc["advertencias_periodo"]:
            ws.append([w])

    # ---------- 8. Deductibilidad — Por empleado ----------
    if deduc.get("desglose_por_empleado"):
        ws = wb.create_sheet("Deductibilidad — Por empleado")
        ws.freeze_panes = "A2"
        cols = [
            "RFC", "Nombre", "Periodicidad", "Periodos detectados",
            "Meses detectados", "Percepciones gravadas",
            "ISR retenido", "ISR teórico", "Diferencia", "Advertencia",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for emp in deduc["desglose_por_empleado"]:
            ws.append([
                emp["rfc"], emp["nombre"], emp["periodicidad"],
                emp["periodos_detectados"], emp.get("meses_detectados"),
                emp["percepciones_gravadas"], emp["isr_retenido"],
                emp["isr_teorico"], emp["diferencia"],
                emp.get("advertencia_periodo") or "",
            ])

    # ---------- 9. IMSS ----------
    if imss.get("registros"):
        ws = wb.create_sheet("IMSS")
        ws.freeze_panes = "A2"
        cols = [
            "RFC", "NSS", "Nombre", "Tipo régimen", "Riesgo de trabajo",
            "Fechas registradas", "Días trabajados",
            "SBC", "SDI",
            "Aportaciones patrón", "Aportaciones obrero", "Seguro social retenido",
            "Observaciones",
        ]
        ws.append(_header_row(ws, cols, header_font, header_fill))
        for r in imss["registros"]:
            ws.append([
                r["rfc"], r["nss"], r["nombre"], r["tipo_regimen"], r["riesgo_trabajo"],
                " | ".join(r["fechas_registro"]), r["dias_trabajados"],
                r["salario_base_cot_apor"], r["salario_diario_integrado"],
                r["aportaciones_patronal"], r["aportaciones_obrero"], r["seguro_social"],
                " | ".join(r["observaciones"]),
            ])

    # ---------- 10. IMSS — Alertas (condicional) ----------
    alertas = imss.get("alertas") or {}
    hay_alertas = any(alertas.get(k) for k in ("empleados_sin_nss", "sbc_fuera_limites", "dias_anomalous"))
    if hay_alertas:
        ws = wb.create_sheet("IMSS — Alertas")
        ws.freeze_panes = "A2"
        ws.append(_header_row(ws, ["Tipo de alerta", "RFC"], alert_font, alert_fill))
        for rfc in alertas.get("empleados_sin_nss") or []:
            ws.append(["Sin NSS", rfc])
        for rfc in alertas.get("sbc_fuera_limites") or []:
            ws.append(["SBC fuera de límites", rfc])
        for rfc in alertas.get("dias_anomalous") or []:
            ws.append(["Días anómalos (>31)", rfc])

    # ---------- 11. Periodo vs Periodo (condicional) ----------
    if pvsp and not pvsp.get("insuficiente"):
        ws = wb.create_sheet("Periodo vs Periodo")
        ws.freeze_panes = "A2"
        ws.append(_header_row(ws, ["Métrica", "Previo", "Actual", "Variación", "%"], header_font, header_fill))
        prev = pvsp["periodo_previo"]
        act = pvsp["periodo_actual"]
        v = pvsp["variaciones"]
        ws.append(["Inicio", prev["inicio"], act["inicio"], "", ""])
        ws.append(["Fin", prev["fin"], act["fin"], "", ""])
        ws.append([
            "Empleados",
            prev["total_empleados"], act["total_empleados"],
            v["empleados_variacion"], f"{v['empleados_variacion_pct']:.1f}%",
        ])
        ws.append([
            "Percepciones",
            prev["total_percepciones"], act["total_percepciones"],
            v["percepciones_variacion"], f"{v['percepciones_variacion_pct']:.1f}%",
        ])
        ws.append([
            "Deducciones",
            prev["total_deducciones"], act["total_deducciones"],
            v["deducciones_variacion"], f"{v['deducciones_variacion_pct']:.1f}%",
        ])
        ws.append([
            "Promedio por empleado",
            prev["promedio_por_empleado"], act["promedio_por_empleado"],
            "", "",
        ])
        ws.append([])
        ws.append(["Tendencia", pvsp["analisis_detallado"]["tendencia"]])
        for obs in pvsp["analisis_detallado"]["observaciones"]:
            ws.append(["Observación", obs])
        if v["empleados_nuevos"]:
            ws.append([])
            ws.append(["Empleados nuevos"])
            for rfc in v["empleados_nuevos"]:
                ws.append([rfc])
        if v["empleados_eliminados"]:
            ws.append([])
            ws.append(["Empleados eliminados"])
            for rfc in v["empleados_eliminados"]:
                ws.append([rfc])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
