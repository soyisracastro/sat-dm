"""Tests del procesador de Nómina (parser, persistencia, reportes, export)."""

from __future__ import annotations

import io
import json

import pytest

from sat_descarga.procesador import (
    CfdiData,
    ConceptoNomina,
    DatosNomina,
    ProcesadorDB,
    parse_cfdi,
)
from sat_descarga.procesador import db as db_mod
from sat_descarga.procesador.constants_nomina import (
    calcular_isr_bruto,
    calcular_spe,
    get_isr_tarifa,
    get_tarifa_year_label,
)
from sat_descarga.procesador.exportar_nomina import to_xlsx
from sat_descarga.procesador.reportes_nomina import (
    conceptos_de_recibo,
    listar_recibos,
    reporte_deducibilidad,
    reporte_imss,
    reporte_periodo_vs_periodo,
    stats_nomina,
)

# Empresa dueña del buffer en los tests (el receptor de los fixtures).
MI_RFC = "BBB020202BBB"
F_RFC = {"mi_rfc": MI_RFC}


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _recibo(
    uuid: str,
    *,
    rfc: str = "TRABA800101AAA",
    nombre: str = "Juan Pérez",
    nss: str = "12345678901",
    sbc: float = 500.0,
    sdi: float = 525.0,
    fecha_pago: str = "2026-01-31",
    periodicidad: str = "05",         # mensual
    tipo_nomina: str = "O",
    dias: float = 30.0,
    percepciones: list[tuple[str, float, float]] | None = None,  # (tipo, gravado, exento)
    deducciones: list[tuple[str, float]] | None = None,           # (tipo, importe)
    otros_pagos: list[tuple[str, float, float]] | None = None,    # (tipo, importe, subsidio_causado)
) -> CfdiData:
    """
    Crea un CFDI tipo N sintético con su complemento de nómina.

    Por default es un recibo "salario base" de un mes completo, $20,000 gravados
    ($666.67/día × 30 días). Override con kwargs según el test.
    """
    if percepciones is None:
        percepciones = [("001", 20000.0, 0.0)]
    if deducciones is None:
        deducciones = [("001", 1500.0), ("002", 2500.0)]
    if otros_pagos is None:
        otros_pagos = []

    conceptos: list[ConceptoNomina] = []
    total_perc = 0.0
    for tipo, gravado, exento in percepciones:
        conceptos.append(
            ConceptoNomina(
                clase="Percepcion",
                tipo_concepto=tipo,
                concepto=f"Percepción {tipo}",
                importe_gravado=gravado,
                importe_exento=exento,
            )
        )
        total_perc += gravado + exento

    total_ded = 0.0
    for tipo, importe in deducciones:
        conceptos.append(
            ConceptoNomina(
                clase="Deduccion",
                tipo_concepto=tipo,
                concepto=f"Deducción {tipo}",
                importe=importe,
            )
        )
        total_ded += importe

    total_otros = 0.0
    for tipo, importe, subsidio in otros_pagos:
        conceptos.append(
            ConceptoNomina(
                clase="OtroPago",
                tipo_concepto=tipo,
                concepto=f"Otro pago {tipo}",
                importe=importe,
                subsidio_causado=subsidio,
            )
        )
        total_otros += importe

    datos_nomina = DatosNomina(
        registro_patronal="A1234567890",
        curp="PEPJ800101HDFRJN09",
        nss=nss,
        num_empleado="001",
        puesto="Desarrollador",
        tipo_contrato="01",
        tipo_regimen="02",
        periodicidad_pago=periodicidad,
        salario_base_cot_apor=sbc,
        salario_diario_integrado=sdi,
        riesgo_trabajo="1",
        tipo_nomina=tipo_nomina,
        fecha_pago=fecha_pago,
        fecha_inicial_pago=fecha_pago,
        fecha_final_pago=fecha_pago,
        num_dias_pagados=dias,
        total_percepciones=total_perc,
        total_deducciones=total_ded,
        total_otros_pagos=total_otros,
        conceptos=conceptos,
    )

    cfdi = CfdiData(
        uuid=uuid,
        version="4.0",
        tipo_comprobante="N",
        fecha_emision=fecha_pago + "T10:00:00",
        emisor_rfc="PAT900101AAA",
        emisor_nombre="Patrón S.A.",
        receptor_rfc=rfc,
        receptor_nombre=nombre,
        sub_total=total_perc,
        total=total_perc - total_ded + total_otros,
        moneda="MXN",
    )
    cfdi.datos_nomina = datos_nomina
    return cfdi


def _ingreso(uuid: str = "INGRESO1-1111-1111-1111-111111111111") -> CfdiData:
    """CFDI tipo I (para test 'no parsea nomina cuando tipo no es N')."""
    return CfdiData(
        uuid=uuid,
        version="4.0",
        tipo_comprobante="I",
        fecha_emision="2026-01-15T10:00:00",
        emisor_rfc="AAA010101AAA",
        emisor_nombre="Emisor",
        receptor_rfc="BBB020202BBB",
        receptor_nombre="Receptor",
        sub_total=1000.0,
        total=1160.0,
    )


@pytest.fixture()
def db(tmp_path):
    db_mod.resetear_singleton_para_tests()
    inst = ProcesadorDB(tmp_path / "test.db")
    yield inst
    inst.close()
    db_mod.resetear_singleton_para_tests()


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


_XML_NOMINA = """<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
                  xmlns:nomina12="http://www.sat.gob.mx/nomina12"
                  Version="4.0" TipoDeComprobante="N"
                  Fecha="2026-01-31T10:00:00" SubTotal="20000" Total="16000"
                  Moneda="MXN">
  <cfdi:Emisor Rfc="PAT900101AAA" Nombre="Patrón SA" RegimenFiscal="601"/>
  <cfdi:Receptor Rfc="TRABA800101AAA" Nombre="Juan Pérez" UsoCFDI="CN01"/>
  <cfdi:Complemento>
    <nomina12:Nomina TipoNomina="O" FechaPago="2026-01-31"
                     FechaInicialPago="2026-01-01" FechaFinalPago="2026-01-31"
                     NumDiasPagados="30" TotalPercepciones="20500"
                     TotalDeducciones="4000" TotalOtrosPagos="500">
      <nomina12:Emisor RegistroPatronal="A1234567890"/>
      <nomina12:Receptor Curp="PEPJ800101HDFRJN09" NumSeguridadSocial="12345678901"
                        NumEmpleado="001" Puesto="Dev" TipoContrato="01"
                        TipoRegimen="02" PeriodicidadPago="05"
                        SalarioBaseCotApor="500" SalarioDiarioIntegrado="525"
                        RiesgoTrabajo="1" ClaveEntFed="09"/>
      <nomina12:Percepciones>
        <nomina12:Percepcion TipoPercepcion="001" Clave="P001"
                             Concepto="Sueldo" ImporteGravado="20000"
                             ImporteExento="0"/>
        <nomina12:Percepcion TipoPercepcion="021" Clave="P021"
                             Concepto="Prima vacacional" ImporteGravado="0"
                             ImporteExento="500"/>
      </nomina12:Percepciones>
      <nomina12:Deducciones>
        <nomina12:Deduccion TipoDeduccion="002" Clave="D002"
                            Concepto="ISR" Importe="2500"/>
      </nomina12:Deducciones>
      <nomina12:OtrosPagos>
        <nomina12:OtroPago TipoOtroPago="002" Clave="O002"
                          Concepto="Subsidio al empleo" Importe="500">
          <nomina12:SubsidioAlEmpleo SubsidioCausado="500"/>
        </nomina12:OtroPago>
      </nomina12:OtrosPagos>
    </nomina12:Nomina>
    <tfd:TimbreFiscalDigital xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital"
                             UUID="AAAA1111-2222-3333-4444-555555555555"
                             FechaTimbrado="2026-01-31T11:00:00"/>
  </cfdi:Complemento>
</cfdi:Comprobante>
"""


def test_parser_extrae_nomina_basica():
    """XML real con 2 percepciones + 1 deducción + 1 otro pago se parsea entero."""
    data = parse_cfdi(_XML_NOMINA)
    assert data.tipo_comprobante == "N"
    assert data.datos_nomina is not None

    dn = data.datos_nomina
    assert dn.tipo_nomina == "O"
    assert dn.fecha_pago == "2026-01-31"
    assert dn.registro_patronal == "A1234567890"
    assert dn.nss == "12345678901"
    assert dn.curp == "PEPJ800101HDFRJN09"
    assert dn.periodicidad_pago == "05"
    assert dn.salario_base_cot_apor == 500.0
    assert dn.num_dias_pagados == 30.0
    assert dn.total_percepciones == 20500.0

    assert len(dn.conceptos) == 4
    clases = [c.clase for c in dn.conceptos]
    assert clases.count("Percepcion") == 2
    assert clases.count("Deduccion") == 1
    assert clases.count("OtroPago") == 1


def test_parser_extrae_subsidio_causado_tipo_002():
    """El sub-elemento SubsidioAlEmpleo del OtroPago 002 se lee correctamente."""
    data = parse_cfdi(_XML_NOMINA)
    otros = [c for c in data.datos_nomina.conceptos if c.clase == "OtroPago"]
    assert len(otros) == 1
    assert otros[0].tipo_concepto == "002"
    assert otros[0].subsidio_causado == 500.0


def test_parser_no_extrae_nomina_si_tipo_no_es_n():
    """Un CFDI tipo I no debe llenar datos_nomina."""
    xml_i = _XML_NOMINA.replace('TipoDeComprobante="N"', 'TipoDeComprobante="I"')
    data = parse_cfdi(xml_i)
    assert data.datos_nomina is None


# ---------------------------------------------------------------------------
# Persistencia (db.agregar + repoblación)
# ---------------------------------------------------------------------------


def test_db_agregar_inserta_recibo_y_conceptos(db):
    cfdi = _recibo("NOM00001-1111-1111-1111-111111111111")
    db.agregar([cfdi], mi_rfc=MI_RFC)

    with db.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM nomina_recibos")
        assert cur.fetchone()[0] == 1
        cur.execute("SELECT COUNT(*) FROM nomina_conceptos")
        # 1 percepción + 2 deducciones = 3
        assert cur.fetchone()[0] == 3
        cur.execute(
            "SELECT nss, salario_base_cot_apor, total_percepciones "
            "FROM nomina_recibos WHERE cfdi_uuid = ?",
            (cfdi.uuid,),
        )
        r = cur.fetchone()
        assert r["nss"] == "12345678901"
        assert r["salario_base_cot_apor"] == 500.0


def test_db_borrar_limpia_nomina_tables(db):
    db.agregar([_recibo("NOM00001-1111-1111-1111-111111111111")], mi_rfc=MI_RFC)
    db.borrar(MI_RFC)
    with db.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM nomina_recibos")
        assert cur.fetchone()[0] == 0
        cur.execute("SELECT COUNT(*) FROM nomina_conceptos")
        assert cur.fetchone()[0] == 0


def test_db_repoblar_nomina_idempotente(db):
    """Llamar _repoblar_nomina dos veces no debe duplicar filas."""
    cfdi = _recibo("NOM00001-1111-1111-1111-111111111111")
    db.agregar([cfdi], mi_rfc=MI_RFC)

    # Vaciar las tablas de nómina pero dejar el CFDI en `cfdis` con su raw_json.
    with db._lock, db._conn:
        db._conn.execute("DELETE FROM nomina_conceptos")
        db._conn.execute("DELETE FROM nomina_recibos")

    db._repoblar_nomina()
    with db.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM nomina_recibos")
        assert cur.fetchone()[0] == 1

    # Segunda repoblación: no debe agregar nuevamente.
    db._repoblar_nomina()
    with db.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM nomina_recibos")
        assert cur.fetchone()[0] == 1


# ---------------------------------------------------------------------------
# Stats y listar
# ---------------------------------------------------------------------------


def test_stats_nomina(db):
    db.agregar([
        _recibo("NOM00001-1111-1111-1111-111111111111", rfc="AAA111", nombre="A"),
        _recibo("NOM00002-2222-2222-2222-222222222222", rfc="AAA111", nombre="A"),
        _recibo("NOM00003-3333-3333-3333-333333333333", rfc="BBB222", nombre="B"),
    ], mi_rfc=MI_RFC)
    s = stats_nomina(db, F_RFC)
    assert s["total_recibos"] == 3
    assert s["total_empleados"] == 2
    assert s["nominas_ordinarias"] == 3
    # Cada recibo default: 1 percep + 2 deduc → 3×3 conceptos
    assert s["total_conceptos"] == 9
    assert s["total_global_recibos"] == 3


def test_listar_recibos_filtros(db):
    db.agregar([
        _recibo("NOM00001-1111-1111-1111-111111111111", tipo_nomina="O",
                periodicidad="04", fecha_pago="2026-01-15"),
        _recibo("NOM00002-2222-2222-2222-222222222222", tipo_nomina="E",
                periodicidad="05", fecha_pago="2026-02-15"),
    ], mi_rfc=MI_RFC)

    todos = listar_recibos(db)
    assert todos["total"] == 2

    solo_ord = listar_recibos(db, {"tipo_nomina": "O"})
    assert solo_ord["total"] == 1
    assert solo_ord["items"][0]["cfdi_uuid"].startswith("NOM00001")

    solo_quincenal = listar_recibos(db, {"periodicidad": "04"})
    assert solo_quincenal["total"] == 1

    por_fecha = listar_recibos(db, {"desde": "2026-02-01", "hasta": "2026-02-28"})
    assert por_fecha["total"] == 1
    assert por_fecha["items"][0]["cfdi_uuid"].startswith("NOM00002")


def test_conceptos_de_recibo_orden(db):
    db.agregar([_recibo("NOM00001-1111-1111-1111-111111111111")], mi_rfc=MI_RFC)
    conceptos = conceptos_de_recibo(db, "NOM00001-1111-1111-1111-111111111111", MI_RFC)
    # Default: 1 percep + 2 deduc → 3 conceptos
    assert len(conceptos) == 3
    # Percepciones primero, luego Deducciones, luego OtrosPagos.
    assert conceptos[0]["clase"] == "Percepcion"
    assert conceptos[1]["clase"] == "Deduccion"
    assert conceptos[2]["clase"] == "Deduccion"


# ---------------------------------------------------------------------------
# Reporte Deductibilidad
# ---------------------------------------------------------------------------


def test_reporte_deducibilidad_calculo_isr_2026(db):
    """
    Empleado mensual con $20,000 gravados en 2026 → ISR bruto e ISR teórico
    se calculan con la tarifa 2026 y sin SPE (gravado > limite_spe).
    """
    db.agregar([_recibo(
        "NOM00001-1111-1111-1111-111111111111",
        percepciones=[("001", 20000.0, 0.0)],
        deducciones=[("002", 2500.0)],
        fecha_pago="2026-01-31",
    )], mi_rfc=MI_RFC)
    r = reporte_deducibilidad(db)
    isr = r["isr_analisis"]
    assert isr["year_detected"] == 2026
    assert isr["tarifa_label"] == "2026"

    isr_esperado = calcular_isr_bruto(20000.0, 2026)
    assert isr["isr_bruto"] == pytest.approx(isr_esperado, abs=0.01)
    # $20,000 > 11492 → no aplica SPE
    assert isr["subsidio_aplicado"] == 0.0
    assert isr["isr_teorico"] == pytest.approx(isr_esperado, abs=0.01)
    assert r["isr_retenido"] == 2500.0
    assert r["empleados_analizados"] == 1


def test_reporte_deducibilidad_multiples_meses(db):
    """Mismo empleado con 2 meses cargados → ISR teórico es suma mensual."""
    db.agregar([
        _recibo("NOM00001-1111-1111-1111-111111111111",
                percepciones=[("001", 20000.0, 0.0)],
                fecha_pago="2026-01-31"),
        _recibo("NOM00002-2222-2222-2222-222222222222",
                percepciones=[("001", 20000.0, 0.0)],
                fecha_pago="2026-02-28"),
    ], mi_rfc=MI_RFC)
    r = reporte_deducibilidad(db)
    isr_un_mes = calcular_isr_bruto(20000.0, 2026)
    # ISR teórico total = ISR mensual × 2.
    assert r["isr_analisis"]["isr_bruto"] == pytest.approx(isr_un_mes * 2, abs=0.02)
    # Y el empleado tiene meses_detectados=2.
    emp = r["desglose_por_empleado"][0]
    assert emp["meses_detectados"] == 2


def test_reporte_deducibilidad_periodo_incompleto_advertencia(db):
    """Empleado quincenal con solo 1 recibo del mes → advertencia poblada."""
    db.agregar([_recibo(
        "NOM00001-1111-1111-1111-111111111111",
        periodicidad="04",      # quincenal → se esperan 2 por mes
        fecha_pago="2026-01-15",
    )], mi_rfc=MI_RFC)
    r = reporte_deducibilidad(db)
    emp = r["desglose_por_empleado"][0]
    assert emp["advertencia_periodo"] is not None
    assert "incompleto" in emp["advertencia_periodo"].lower()
    assert r["advertencias_periodo"] != []


# ---------------------------------------------------------------------------
# Reporte IMSS
# ---------------------------------------------------------------------------


def test_reporte_imss_aportaciones(db):
    """SBC=$500 → patronal 500×0.1875=93.75, obrero 500×0.0625=31.25."""
    db.agregar([_recibo(
        "NOM00001-1111-1111-1111-111111111111",
        sbc=500.0,
    )], mi_rfc=MI_RFC)
    r = reporte_imss(db)
    assert r["total_empleados"] == 1
    reg = r["registros"][0]
    assert reg["aportaciones_patronal"] == pytest.approx(93.75, abs=0.01)
    assert reg["aportaciones_obrero"] == pytest.approx(31.25, abs=0.01)


def test_reporte_imss_alertas_sin_nss(db):
    db.agregar([_recibo(
        "NOM00001-1111-1111-1111-111111111111",
        rfc="AAA111",
        nss="",
    )], mi_rfc=MI_RFC)
    r = reporte_imss(db)
    assert "AAA111" in r["alertas"]["empleados_sin_nss"]


def test_reporte_imss_alertas_sbc_fuera(db):
    db.agregar([_recibo(
        "NOM00001-1111-1111-1111-111111111111",
        rfc="AAA111",
        sbc=999_999.0,
    )], mi_rfc=MI_RFC)
    r = reporte_imss(db)
    assert "AAA111" in r["alertas"]["sbc_fuera_limites"]


# ---------------------------------------------------------------------------
# Reporte Periodo vs Periodo
# ---------------------------------------------------------------------------


def test_reporte_periodo_vs_periodo_insuficiente(db):
    db.agregar([_recibo("NOM00001-1111-1111-1111-111111111111", fecha_pago="2026-01-31")], mi_rfc=MI_RFC)
    r = reporte_periodo_vs_periodo(db)
    assert r["insuficiente"] is True
    assert "dos meses" in (r["mensaje_insuficiente"] or "").lower()


def test_reporte_periodo_vs_periodo_variaciones(db):
    """2 meses con cambio de plantilla → empleados_nuevos/eliminados detectados."""
    db.agregar([
        # Mes previo: empleado A
        _recibo("NOM00001-1111-1111-1111-111111111111", rfc="AAA111",
                fecha_pago="2026-01-31"),
        # Mes actual: empleado B (A salió, B entró)
        _recibo("NOM00002-2222-2222-2222-222222222222", rfc="BBB222",
                fecha_pago="2026-02-28"),
    ], mi_rfc=MI_RFC)
    r = reporte_periodo_vs_periodo(db)
    assert r["insuficiente"] is False
    assert "AAA111" in r["variaciones"]["empleados_eliminados"]
    assert "BBB222" in r["variaciones"]["empleados_nuevos"]


# ---------------------------------------------------------------------------
# Export XLSX
# ---------------------------------------------------------------------------


def test_exportar_nomina_xlsx_estructura(db):
    db.agregar([_recibo("NOM00001-1111-1111-1111-111111111111")], mi_rfc=MI_RFC)
    xlsx_bytes = to_xlsx(db, F_RFC)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    # Sheets siempre presentes
    assert "Disclaimer" in wb.sheetnames
    assert "Resumen" in wb.sheetnames
    assert "Recibos" in wb.sheetnames
    assert "Percepciones" in wb.sheetnames
    assert "Deducciones" in wb.sheetnames
    # Deductibilidad global siempre
    assert "Deductibilidad — Global" in wb.sheetnames
    # IMSS siempre que hay recibos
    assert "IMSS" in wb.sheetnames


def test_exportar_nomina_xlsx_disclaimer_first(db):
    db.agregar([_recibo("NOM00001-1111-1111-1111-111111111111")], mi_rfc=MI_RFC)
    xlsx_bytes = to_xlsx(db, F_RFC)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    assert wb.sheetnames[0] == "Disclaimer"


# ---------------------------------------------------------------------------
# Constantes ISR + SPE
# ---------------------------------------------------------------------------


def test_isr_tarifa_year_selection():
    """get_isr_tarifa selecciona la tarifa correcta por año."""
    # 2021 → tarifa 2020-2022 (primer rango: 0.01 - 644.58)
    t_2021 = get_isr_tarifa(2021)
    assert t_2021[0][1] == 644.58
    # 2024 → tarifa 2023-2025 (primer rango: 0.01 - 746.04)
    t_2024 = get_isr_tarifa(2024)
    assert t_2024[0][1] == 746.04
    # 2026 → tarifa 2026 (primer rango: 0.01 - 844.59)
    t_2026 = get_isr_tarifa(2026)
    assert t_2026[0][1] == 844.59
    # Etiquetas
    assert get_tarifa_year_label(2021) == "2020-2022"
    assert get_tarifa_year_label(2024) == "2023-2025"
    assert get_tarifa_year_label(2026) == "2026"


def test_spe_2026_enero_vs_resto():
    """En 2026 el SPE de enero es distinto del SPE del resto del año."""
    enero = calcular_spe(8000.0, 2026, mes=1)
    resto = calcular_spe(8000.0, 2026, mes=2)
    assert enero != resto
    # Para ingreso $8000 < $11492.66 (límite SPE 2026), ambos deben ser > 0.
    assert enero > 0
    assert resto > 0


# ---------------------------------------------------------------------------
# Filtros persistidos con key separada
# ---------------------------------------------------------------------------


def test_filtros_nomina_separados_de_cfdi_y_pagos(db):
    """`nomina_actuales` y `actuales` son keys independientes."""
    db.filtros_set({"tipo": "I"}, key="actuales")
    db.filtros_set({"status": ["pago_parcial"]}, key="pagos_actuales")
    db.filtros_set({"tipo_nomina": "O"}, key="nomina_actuales")

    assert db.filtros_get(key="actuales") == {"tipo": "I"}
    assert db.filtros_get(key="pagos_actuales") == {"status": ["pago_parcial"]}
    assert db.filtros_get(key="nomina_actuales") == {"tipo_nomina": "O"}
