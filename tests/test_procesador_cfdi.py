"""Tests del procesador de comprobantes (parser, validaciones, DB, reportes, export)."""

import io
from pathlib import Path

import pytest

from sat_descarga.procesador import (
    CfdiData,
    ConceptoCfdi,
    ProcesadorDB,
    parse_cfdi,
)
from sat_descarga.procesador import db as db_mod
from sat_descarga.procesador.cfdi_parser import CfdiParseError
from sat_descarga.procesador.exportar import to_csv, to_xlsx
from sat_descarga.procesador.reportes_cfdi import (
    integridad,
    stats_generales,
    top_contrapartes,
    totales_por_mes,
)
from sat_descarga.procesador.validaciones import validar, validar_y_anotar


# ---------------------------------------------------------------------------
# Fixtures: XMLs sintéticos representativos
# ---------------------------------------------------------------------------


def _cfdi_ingreso_xml(uuid: str = "AAAAAAAA-1111-2222-3333-444444444444",
                     fecha: str = "2026-05-15T10:00:00",
                     total: float = 1160.0,
                     sub_total: float = 1000.0,
                     iva: float = 160.0,
                     emisor_rfc: str = "AAA010101AAA",
                     emisor_nombre: str = "EMPRESA EMISORA SA DE CV",
                     receptor_rfc: str = "BBB020202BBB",
                     receptor_nombre: str = "CLIENTE RECEPTOR") -> bytes:
    """CFDI 4.0 de Ingreso con 1 concepto y un IVA 16% trasladado."""
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante
    xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
    Version="4.0" Serie="A" Folio="100"
    Fecha="{fecha}" Sello="..." NoCertificado="000" Certificado="..."
    SubTotal="{sub_total}" Moneda="MXN" Total="{total}"
    TipoDeComprobante="I" Exportacion="01" MetodoPago="PUE" FormaPago="03"
    LugarExpedicion="64000">
  <cfdi:Emisor Rfc="{emisor_rfc}" Nombre="{emisor_nombre}" RegimenFiscal="601"/>
  <cfdi:Receptor Rfc="{receptor_rfc}" Nombre="{receptor_nombre}"
                 DomicilioFiscalReceptor="64000" RegimenFiscalReceptor="612"
                 UsoCFDI="G03"/>
  <cfdi:Conceptos>
    <cfdi:Concepto ClaveProdServ="01010101" Cantidad="1" ClaveUnidad="ACT"
                   Descripcion="Servicio profesional"
                   ValorUnitario="{sub_total}" Importe="{sub_total}"
                   ObjetoImp="02">
      <cfdi:Impuestos>
        <cfdi:Traslados>
          <cfdi:Traslado Base="{sub_total}" Impuesto="002" TipoFactor="Tasa"
                         TasaOCuota="0.160000" Importe="{iva}"/>
        </cfdi:Traslados>
      </cfdi:Impuestos>
    </cfdi:Concepto>
  </cfdi:Conceptos>
  <cfdi:Impuestos TotalImpuestosTrasladados="{iva}">
    <cfdi:Traslados>
      <cfdi:Traslado Base="{sub_total}" Impuesto="002" TipoFactor="Tasa"
                     TasaOCuota="0.160000" Importe="{iva}"/>
    </cfdi:Traslados>
  </cfdi:Impuestos>
  <cfdi:Complemento>
    <tfd:TimbreFiscalDigital
        xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital"
        Version="1.1" UUID="{uuid}"
        FechaTimbrado="2026-05-15T10:01:00"
        RfcProvCertif="SAT970701NN3"
        SelloCFD="..." SelloSAT="..." NoCertificadoSAT="000"/>
  </cfdi:Complemento>
</cfdi:Comprobante>
""".encode("utf-8")


def _cfdi_pago_xml(uuid: str = "PPPPPPPP-1111-2222-3333-444444444444") -> bytes:
    """CFDI 4.0 tipo P con complemento Pagos 2.0 y 1 doctoRelacionado."""
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante
    xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
    Version="4.0" Serie="P" Folio="50"
    Fecha="2026-06-01T12:00:00" Sello="..." NoCertificado="000" Certificado="..."
    SubTotal="0" Moneda="XXX" Total="0"
    TipoDeComprobante="P" Exportacion="01"
    LugarExpedicion="64000">
  <cfdi:Emisor Rfc="AAA010101AAA" Nombre="EMPRESA EMISORA SA DE CV" RegimenFiscal="601"/>
  <cfdi:Receptor Rfc="BBB020202BBB" Nombre="CLIENTE RECEPTOR"
                 DomicilioFiscalReceptor="64000" RegimenFiscalReceptor="612"
                 UsoCFDI="CP01"/>
  <cfdi:Conceptos>
    <cfdi:Concepto ClaveProdServ="84111506" Cantidad="1" ClaveUnidad="ACT"
                   Descripcion="Pago" ValorUnitario="0" Importe="0" ObjetoImp="01"/>
  </cfdi:Conceptos>
  <cfdi:Complemento>
    <pago20:Pagos xmlns:pago20="http://www.sat.gob.mx/Pagos20" Version="2.0">
      <pago20:Pago FechaPago="2026-06-01T11:00:00" FormaDePagoP="03"
                   MonedaP="MXN" Monto="1160.00">
        <pago20:DoctoRelacionado IdDocumento="AAAAAAAA-1111-2222-3333-444444444444"
                                 Serie="A" Folio="100" MonedaDR="MXN"
                                 MetodoDePagoDR="PPD" NumParcialidad="1"
                                 ImpSaldoAnt="1160.00" ImpPagado="1160.00"
                                 ImpSaldoInsoluto="0.00" ObjetoImpDR="01"/>
      </pago20:Pago>
    </pago20:Pagos>
    <tfd:TimbreFiscalDigital
        xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital"
        Version="1.1" UUID="{uuid}"
        FechaTimbrado="2026-06-01T12:01:00"
        RfcProvCertif="SAT970701NN3"
        SelloCFD="..." SelloSAT="..." NoCertificadoSAT="000"/>
  </cfdi:Complemento>
</cfdi:Comprobante>
""".encode("utf-8")


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def test_parse_cfdi_ingreso_extrae_cabecera_y_montos():
    cfdi = parse_cfdi(_cfdi_ingreso_xml())
    assert cfdi.uuid == "AAAAAAAA-1111-2222-3333-444444444444"
    assert cfdi.version == "4.0"
    assert cfdi.tipo_comprobante == "I"
    assert cfdi.serie == "A"
    assert cfdi.folio == "100"
    assert cfdi.emisor_rfc == "AAA010101AAA"
    assert cfdi.receptor_rfc == "BBB020202BBB"
    assert cfdi.receptor_uso_cfdi == "G03"
    assert cfdi.sub_total == 1000.0
    assert cfdi.total == 1160.0
    assert cfdi.iva_trasladado == 160.0
    assert cfdi.iva_retenido == 0.0
    assert cfdi.isr_retenido == 0.0
    assert cfdi.moneda == "MXN"
    assert len(cfdi.conceptos) == 1
    assert cfdi.conceptos[0].descripcion == "Servicio profesional"


def test_parse_cfdi_tipo_p_extrae_datos_pago():
    cfdi = parse_cfdi(_cfdi_pago_xml())
    assert cfdi.tipo_comprobante == "P"
    assert cfdi.datos_pago is not None
    assert cfdi.datos_pago.monto_pago == 1160.0
    assert cfdi.datos_pago.forma_de_pago == "03"
    assert len(cfdi.datos_pago.documentos_relacionados) == 1
    dr = cfdi.datos_pago.documentos_relacionados[0]
    assert dr.id_documento == "AAAAAAAA-1111-2222-3333-444444444444"
    assert dr.metodo_de_pago_dr == "PPD"
    assert dr.imp_pagado == 1160.0


def test_parse_cfdi_xml_invalido():
    with pytest.raises(CfdiParseError):
        parse_cfdi(b"<NoSoyUnCfdi/>")


# ---------------------------------------------------------------------------
# Validaciones
# ---------------------------------------------------------------------------


def test_validar_cfdi_coherente_sin_warnings():
    cfdi = parse_cfdi(_cfdi_ingreso_xml())
    assert validar(cfdi) == []


def test_validar_monto_incoherente_levanta_warning():
    cfdi = parse_cfdi(_cfdi_ingreso_xml(total=2000.0))  # 2000 ≠ 1000 + 160
    warnings = validar(cfdi)
    assert any("Total" in w for w in warnings)


def test_validar_y_anotar_pobla_warnings():
    cfdi = parse_cfdi(_cfdi_ingreso_xml(total=2000.0))
    validar_y_anotar(cfdi)
    assert cfdi.warnings != []


def test_parse_extrae_ieps_y_valida_integridad():
    """CFDI con IEPS trasladado: el parser lo extrae y la validación lo suma
    correctamente sin levantar warnings (caso telecomunicaciones)."""
    xml = """<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante
    xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
    Version="4.0" Serie="A" Folio="100"
    Fecha="2026-05-15T10:00:00" Sello="..." NoCertificado="000" Certificado="..."
    SubTotal="754.62" Descuento="80.00" Total="798.99"
    TipoDeComprobante="I" Exportacion="01" MetodoPago="PPD" FormaPago="99"
    LugarExpedicion="06500" Moneda="MXN">
  <cfdi:Emisor Rfc="TME840315KT6" Nombre="TELEFONOS DE MEXICO" RegimenFiscal="601"/>
  <cfdi:Receptor Rfc="GUPF620405TD8" Nombre="FERNANDO GUZMAN PINEDA"
                 DomicilioFiscalReceptor="40660" RegimenFiscalReceptor="612"
                 UsoCFDI="G03"/>
  <cfdi:Conceptos>
    <cfdi:Concepto ClaveProdServ="81161700" Cantidad="1" ClaveUnidad="E48"
                   Descripcion="Servicios de Telecomunicaciones"
                   ValorUnitario="754.62" Importe="754.62" Descuento="80.00"
                   ObjetoImp="02"/>
  </cfdi:Conceptos>
  <cfdi:Impuestos TotalImpuestosTrasladados="124.37">
    <cfdi:Traslados>
      <cfdi:Traslado Base="688.79" Importe="110.20" Impuesto="002"
                     TasaOCuota="0.160000" TipoFactor="Tasa"/>
      <cfdi:Traslado Base="472.28" Importe="14.17" Impuesto="003"
                     TasaOCuota="0.030000" TipoFactor="Tasa"/>
    </cfdi:Traslados>
  </cfdi:Impuestos>
  <cfdi:Complemento>
    <tfd:TimbreFiscalDigital
        xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital"
        Version="1.1" UUID="59f11791-008e-4c91-8752-edef6d7a2719"
        FechaTimbrado="2026-05-15T10:01:00"
        RfcProvCertif="TME840315KT6" SelloCFD="..." SelloSAT="..."
        NoCertificadoSAT="000"/>
  </cfdi:Complemento>
</cfdi:Comprobante>
""".encode("utf-8")
    cfdi = parse_cfdi(xml)
    assert cfdi.iva_trasladado == 110.20
    assert cfdi.ieps_trasladado == 14.17
    # Total = 754.62 − 80 + 110.20 + 14.17 = 798.99 → cuadra exacto.
    assert validar(cfdi) == []


# ---------------------------------------------------------------------------
# DB (SQLite en :memory:)
# ---------------------------------------------------------------------------


@pytest.fixture()
def db(tmp_path):
    """DB temporal en tmp_path para no tocar ~/.sat-descarga."""
    db_mod.resetear_singleton_para_tests()
    inst = ProcesadorDB(tmp_path / "test.db")
    yield inst
    inst.close()
    db_mod.resetear_singleton_para_tests()


def test_db_agregar_dedupe_por_uuid(db):
    cfdi1 = validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))
    cfdi2 = validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))  # mismo UUID
    r1 = db.agregar([cfdi1])
    r2 = db.agregar([cfdi2])
    assert r1 == {"agregados": 1, "duplicados": 0}
    assert r2 == {"agregados": 0, "duplicados": 1}
    assert db.count() == 1


def test_db_listar_paginado(db):
    cfdis = []
    for i in range(5):
        xml = _cfdi_ingreso_xml(uuid=f"AAAAAAAA-1111-2222-3333-44444444444{i}")
        cfdis.append(validar_y_anotar(parse_cfdi(xml)))
    db.agregar(cfdis)
    out = db.listar({}, page=1, page_size=3)
    assert out["total"] == 5
    assert len(out["items"]) == 3


def test_db_filtro_por_tipo(db):
    db.agregar([
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml())),
        validar_y_anotar(parse_cfdi(_cfdi_pago_xml())),
    ])
    out_i = db.listar({"tipo": "I"})
    out_p = db.listar({"tipo": "P"})
    assert out_i["total"] == 1
    assert out_p["total"] == 1


def test_db_filtros_get_set_persisten(db):
    db.filtros_set({"desde": "2026-01-01", "tipo": "I", "solo_con_errores": False})
    f = db.filtros_get()
    assert f["desde"] == "2026-01-01"
    assert f["tipo"] == "I"


def test_db_borrar_limpia_todo(db):
    db.agregar([validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))])
    db.filtros_set({"tipo": "I"})
    assert db.count() == 1
    db.borrar()
    assert db.count() == 0
    assert db.filtros_get() == {
        "desde": None, "hasta": None, "tipo": None, "direccion": None,
        "busqueda": None, "solo_con_errores": False,
        "monto_min": None, "monto_max": None,
    }


def test_db_direccion_se_infiere_de_mi_rfc(db):
    """Cuando se pasa mi_rfc, el CFDI se etiqueta E (yo emisor) o R (yo receptor)."""
    cfdi_emi = validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
        uuid="EEEEEEEE-1111-2222-3333-444444444444",
        emisor_rfc="ME01ABC123XYZ", emisor_nombre="Mi Empresa",
    )))
    cfdi_rec = validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
        uuid="RRRRRRRR-1111-2222-3333-444444444444",
        receptor_rfc="ME01ABC123XYZ", receptor_nombre="Mi Empresa",
    )))
    db.agregar([cfdi_emi, cfdi_rec], mi_rfc="ME01ABC123XYZ")

    out_e = db.listar({"direccion": "E"})
    out_r = db.listar({"direccion": "R"})
    assert out_e["total"] == 1
    assert out_e["items"][0]["uuid"] == "EEEEEEEE-1111-2222-3333-444444444444"
    assert out_r["total"] == 1
    assert out_r["items"][0]["uuid"] == "RRRRRRRR-1111-2222-3333-444444444444"


def test_db_direccion_fija_anula_inferencia(db):
    """`direccion_fija` gana sobre cualquier inferencia por mi_rfc."""
    cfdi = validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))
    db.agregar([cfdi], mi_rfc="OTRO123456ABC", direccion_fija="R")
    out = db.listar({"direccion": "R"})
    assert out["total"] == 1


def test_db_actualizar_estado_sat(db):
    db.agregar([validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))])
    uuid = "AAAAAAAA-1111-2222-3333-444444444444"
    db.actualizar_estado_sat(uuid, "Vigente")
    items = db.listar()["items"]
    assert items[0]["estado_sat"] == "Vigente"
    assert items[0]["validado_en"] is not None


# ---------------------------------------------------------------------------
# Reportes SQL
# ---------------------------------------------------------------------------


def test_stats_generales(db):
    db.agregar([
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml())),
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
            uuid="BBBBBBBB-1111-2222-3333-444444444444",
            total=2320.0, sub_total=2000.0, iva=320.0,
        ))),
    ])
    s = stats_generales(db)
    assert s["total_comprobantes"] == 2
    assert s["monto_total"] == 1160.0 + 2320.0
    assert s["iva_trasladado"] == 160.0 + 320.0
    assert s["por_tipo"] == {"I": 2}


def test_totales_por_mes(db):
    db.agregar([
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
            uuid="11111111-1111-2222-3333-444444444444",
            fecha="2026-04-15T10:00:00",
        ))),
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
            uuid="22222222-1111-2222-3333-444444444444",
            fecha="2026-05-15T10:00:00",
        ))),
    ])
    out = totales_por_mes(db)
    meses = {r["mes"] for r in out}
    assert meses == {"2026-04", "2026-05"}


def test_top_contrapartes(db):
    db.agregar([
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
            uuid="11111111-1111-2222-3333-444444444444",
            emisor_rfc="AAA010101AAA", emisor_nombre="EMP1",
        ))),
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
            uuid="22222222-1111-2222-3333-444444444444",
            emisor_rfc="CCC030303CCC", emisor_nombre="EMP2",
            total=5000.0, sub_total=4310.34, iva=689.66,
        ))),
    ])
    out = top_contrapartes(db, n=5)
    # EMP2 tiene más monto → primero
    assert out["emisores"][0]["rfc"] == "CCC030303CCC"
    assert out["emisores"][0]["monto"] == 5000.0


def test_integridad_lista_solo_con_warnings(db):
    db.agregar([
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml())),  # sano
        validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml(
            uuid="22222222-1111-2222-3333-444444444444",
            total=9999.0,  # monto incoherente
        ))),
    ])
    out = integridad(db)
    assert len(out) == 1
    assert out[0]["uuid"] == "22222222-1111-2222-3333-444444444444"


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


def test_to_csv_tiene_bom_y_encabezados(db):
    db.agregar([validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))])
    csv_bytes = to_csv(db)
    # BOM UTF-8 al inicio
    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    text = csv_bytes.decode("utf-8-sig")
    assert "UUID" in text
    assert "AAAAAAAA-1111-2222-3333-444444444444" in text


def test_to_xlsx_genera_dos_sheets(db):
    db.agregar([validar_y_anotar(parse_cfdi(_cfdi_ingreso_xml()))])
    xlsx_bytes = to_xlsx(db)
    # Verificar leyéndolo de vuelta
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    assert "CFDIs" in wb.sheetnames
    assert "Conceptos" in wb.sheetnames
