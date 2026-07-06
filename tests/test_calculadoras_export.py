"""Export de calculadoras a XLSX/PDF y recibos de PTU."""

import io

import pytest

pytest.importorskip("fastapi")
pytest.importorskip("fpdf")
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from sat_descarga.api import server  # noqa: E402
from sat_descarga.cli import config_store  # noqa: E402


@pytest.fixture(autouse=True)
def aislar(tmp_path, monkeypatch):
    monkeypatch.setattr(config_store, "CONFIG_DIR", tmp_path / ".sat-descarga")


@pytest.fixture
def client():
    return TestClient(server.app)


FINIQUITO_INPUTS = {
    "salario": 12000,
    "tipo_salario": "mensual",
    "fecha_ingreso": "2022-03-01",
    "fecha_baja": "2026-07-15",
}

PTU_INPUTS_SIN_EMPRESA = {
    "utilidad_fiscal": 600000,
    "ejercicio": 2025,
    "trabajadores": [
        {
            "nombre": "Trabajador de Prueba",
            "salario_diario": 15000 / 30.4,
            "dias_trabajados": 365,
            "percepcion_anual": 180000,
            "ingreso_mensual_ordinario": 15000,
            "isr_mensual_ordinario": 1068.54,
        }
    ],
}

PTU_INPUTS = {
    "utilidad_fiscal": 600000,
    "ejercicio": 2025,
    "nombre": "Servicios Profesionales del Norte SA de CV",
    "rfc_empresa": "SPN240115ABC",
    "trabajadores": [
        {
            "nombre": "María González Ruiz",
            "rfc": "GORM850412AB1",
            "curp": "GORM850412MDFNZR05",
            "salario_diario": 15000 / 30.4,
            "dias_trabajados": 365,
            "percepcion_anual": 180000,
            "ptu_anio_1": 40000,
            "ptu_anio_2": 40000,
            "ptu_anio_3": 40000,
            "ingreso_mensual_ordinario": 15000,
            "isr_mensual_ordinario": 1068.54,
        },
        {
            "nombre": "José Hernández López",
            "salario_diario": 300,
            "dias_trabajados": 200,
            "percepcion_anual": 60000,
            "ingreso_mensual_ordinario": 9000,
            "isr_mensual_ordinario": 100,
        },
    ],
}


def test_export_xlsx_finiquito(client):
    r = client.post(
        "/calculadoras/exportar/xlsx",
        json={"calculadora": "finiquito", "inputs": FINIQUITO_INPUTS},
    )
    assert r.status_code == 200
    assert r.content[:2] == b"PK"  # magic bytes de ZIP/XLSX
    assert "finiquito_2026.xlsx" in r.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(r.content))
    assert "Resumen" in wb.sheetnames
    assert "Desglose" in wb.sheetnames


def test_export_pdf_finiquito(client):
    r = client.post(
        "/calculadoras/exportar/pdf",
        json={"calculadora": "finiquito", "inputs": FINIQUITO_INPUTS},
    )
    assert r.status_code == 200
    assert r.content[:5] == b"%PDF-"
    assert "finiquito_2026.pdf" in r.headers["content-disposition"]


def test_nombre_archivo_con_rfc(client):
    """Nomenclatura {RFC}_{concepto}_{año}; en PTU el año es el ejercicio."""
    r = client.post(
        "/calculadoras/exportar/xlsx",
        json={"calculadora": "ptu", "inputs": PTU_INPUTS, "rfc": "CAUI890921DAA"},
    )
    assert "CAUI890921DAA_ptu_2025.xlsx" in r.headers["content-disposition"]

    # RFC inválido se omite (no rompe el export)
    r2 = client.post(
        "/calculadoras/exportar/xlsx",
        json={"calculadora": "ptu", "inputs": PTU_INPUTS, "rfc": "no-es-rfc"},
    )
    assert "ptu_2025.xlsx" in r2.headers["content-disposition"]
    assert "no-es-rfc" not in r2.headers["content-disposition"]


@pytest.mark.parametrize(
    "calculadora,inputs",
    [
        ("aguinaldo", {"salario": 12000, "tipo_salario": "mensual", "fecha_ingreso": "2025-01-01", "fecha_calculo": "2026-12-20"}),
        ("sbc", {"salario": 400, "tipo_salario": "diario", "antiguedad_anios": 1}),
        ("isr", {"ingreso_gravado": 15000}),
        (
            "liquidacion",
            {
                "salario": 15000,
                "tipo_salario": "mensual",
                "fecha_ingreso": "2020-03-01",
                "fecha_baja": "2026-07-15",
                "tipo_terminacion": "DESPIDO_INJUSTIFICADO",
            },
        ),
        ("carga-patronal", {"salario": 12000, "tipo_salario": "mensual", "antiguedad_anios": 1}),
        ("ptu", PTU_INPUTS),
    ],
)
def test_export_todas_ambos_formatos(client, calculadora, inputs):
    for formato, magic in (("xlsx", b"PK"), ("pdf", b"%PDF-")):
        r = client.post(
            f"/calculadoras/exportar/{formato}",
            json={"calculadora": calculadora, "inputs": inputs},
        )
        assert r.status_code == 200, f"{calculadora}/{formato}: {r.text[:200]}"
        assert r.content[: len(magic)] == magic


def test_export_ptu_xlsx_incluye_prenomina(client):
    r = client.post(
        "/calculadoras/exportar/xlsx", json={"calculadora": "ptu", "inputs": PTU_INPUTS}
    )
    wb = load_workbook(io.BytesIO(r.content))
    assert "Pre_Nómina" in wb.sheetnames
    ws = wb["Pre_Nómina"]
    # header + 2 trabajadores (más la fila del título)
    headers = [c.value for c in ws[2]]
    assert "Clave perc." in headers


def test_recibos_ptu(client):
    r = client.post(
        "/calculadoras/exportar/recibos-ptu",
        json={"calculadora": "ptu", "inputs": PTU_INPUTS},
    )
    assert r.status_code == 200
    assert r.content[:5] == b"%PDF-"
    assert "recibos-ptu_2025.pdf" in r.headers["content-disposition"]
    # Un recibo (página) por trabajador
    assert r.content.count(b"/Type /Page") >= 2


def test_recibos_solo_para_ptu(client):
    r = client.post(
        "/calculadoras/exportar/recibos-ptu",
        json={"calculadora": "finiquito", "inputs": FINIQUITO_INPUTS},
    )
    assert r.status_code == 400


def test_latin1_translitera_tipografia():
    """Guión largo, puntos suspensivos y comillas tipográficas → ASCII (no «?»)."""
    from sat_descarga.calculadoras.exportar import _latin1

    assert _latin1("PTU — Reparto…") == "PTU - Reparto..."
    assert _latin1("“cita” – ‘x’") == '"cita" - \'x\''
    assert "?" not in _latin1("PTU — Reparto de utilidades · año 2025")


def test_ptu_xlsx_ejercicio_y_empresa_activa(client):
    """El Resumen lleva el ejercicio seleccionado y la empresa activa del catálogo."""
    config_store.add_empresa_ciec("CAUI890921DAA", "Mi Empresa SA de CV", "ciec123")
    r = client.post(
        "/calculadoras/exportar/xlsx",
        json={"calculadora": "ptu", "inputs": PTU_INPUTS_SIN_EMPRESA, "rfc": "CAUI890921DAA"},
    )
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content))["Resumen"]
    assert ws.cell(row=2, column=1).value == "TodoConta · Ejercicio 2025"  # no el año de pago
    filas = {
        ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value
        for i in range(1, 30)
        if ws.cell(row=i, column=1).value
    }
    assert filas["Empresa"] == "Mi Empresa SA de CV"
    assert filas["RFC"] == "CAUI890921DAA"


def test_export_formato_desconocido(client):
    r = client.post(
        "/calculadoras/exportar/docx",
        json={"calculadora": "finiquito", "inputs": FINIQUITO_INPUTS},
    )
    assert r.status_code == 404


def test_export_inputs_invalidos_422(client):
    r = client.post(
        "/calculadoras/exportar/xlsx",
        json={"calculadora": "finiquito", "inputs": {"salario": -1}},
    )
    assert r.status_code == 422
