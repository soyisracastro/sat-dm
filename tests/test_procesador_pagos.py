"""Tests del procesador de Pagos (matcher PPD↔complemento, reportes, export)."""

import io

import pytest

from sat_descarga.procesador import (
    CfdiData,
    ConceptoCfdi,
    ProcesadorDB,
    parse_cfdi,
)
from sat_descarga.procesador import db as db_mod
from sat_descarga.procesador.cfdi_parser import DatosPago, DocumentoRelacionado
from sat_descarga.procesador.exportar_pagos import to_xlsx
from sat_descarga.procesador.reportes_pagos import (
    analisis_fechas,
    detalle_pagos_de_ppd,
    facturas_ppd,
    incidencias_pue,
    pagos_huerfanos,
    stats_pagos,
)
from sat_descarga.procesador.validaciones import validar_y_anotar


# ---------------------------------------------------------------------------
# Fixtures: factories sintéticos para PPD + complemento
# ---------------------------------------------------------------------------


def _ppd(uuid: str, total: float = 1160.0, fecha: str = "2026-01-15T10:00:00",
         emisor_rfc: str = "AAA010101AAA", metodo_pago: str = "PPD") -> CfdiData:
    """CFDI tipo I con metodo_pago=PPD listo para insertar."""
    cfdi = CfdiData(
        uuid=uuid,
        version="4.0",
        tipo_comprobante="I",
        fecha_emision=fecha,
        emisor_rfc=emisor_rfc,
        emisor_nombre="Empresa Emisora",
        receptor_rfc="BBB020202BBB",
        receptor_nombre="Cliente",
        sub_total=1000.0,
        descuento=0.0,
        total=total,
        iva_trasladado=160.0,
        moneda="MXN",
        metodo_pago=metodo_pago,
        forma_pago="99",
    )
    return cfdi


def _complemento(uuid: str, refs: list[tuple[str, float, int]],
                fecha_emision: str = "2026-01-20T10:00:00",
                fecha_pago: str = "2026-01-18T10:00:00",
                metodo_dr: str = "PPD") -> CfdiData:
    """
    CFDI tipo P con N documentos relacionados.
    refs = [(docto_uuid, imp_pagado, num_parcialidad), ...]
    """
    docs = [
        DocumentoRelacionado(
            id_documento=ref_uuid,
            num_parcialidad=parc,
            imp_pagado=imp,
            imp_saldo_ant=imp,
            imp_saldo_insoluto=0.0,
            metodo_de_pago_dr=metodo_dr,
            moneda_dr="MXN",
        )
        for ref_uuid, imp, parc in refs
    ]
    cfdi = CfdiData(
        uuid=uuid,
        version="4.0",
        tipo_comprobante="P",
        fecha_emision=fecha_emision,
        emisor_rfc="AAA010101AAA",
        emisor_nombre="Empresa Emisora",
        receptor_rfc="BBB020202BBB",
        receptor_nombre="Cliente",
        sub_total=0.0,
        total=0.0,
        moneda="XXX",
        datos_pago=DatosPago(
            fecha_pago=fecha_pago,
            forma_de_pago="03",
            moneda_pago="MXN",
            monto_pago=sum(imp for _, imp, _ in refs),
            documentos_relacionados=docs,
        ),
    )
    return cfdi


@pytest.fixture()
def db(tmp_path):
    db_mod.resetear_singleton_para_tests()
    inst = ProcesadorDB(tmp_path / "test.db")
    yield inst
    inst.close()
    db_mod.resetear_singleton_para_tests()


# ---------------------------------------------------------------------------
# Matcher: pagos_relaciones se hidrata al insertar tipo P
# ---------------------------------------------------------------------------


def test_insertar_tipo_p_hidrata_pagos_relaciones(db):
    ppd = _ppd("PPD11111-1111-1111-1111-111111111111")
    pago = _complemento(
        "PAGO1111-1111-1111-1111-111111111111",
        [("PPD11111-1111-1111-1111-111111111111", 1160.0, 1)],
    )
    db.agregar([ppd, pago])

    with db.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM pagos_relaciones")
        assert cur.fetchone()[0] == 1
        cur.execute(
            "SELECT docto_uuid, docto_imp_pagado, docto_num_parcialidad FROM pagos_relaciones"
        )
        r = cur.fetchone()
        assert r["docto_uuid"] == "PPD11111-1111-1111-1111-111111111111"
        assert r["docto_imp_pagado"] == 1160.0
        assert r["docto_num_parcialidad"] == 1


def test_un_complemento_n_documentos(db):
    """Un complemento P puede pagar varias facturas PPD."""
    p1 = _ppd("PPD11111-1111-1111-1111-111111111111", total=500.0)
    p2 = _ppd("PPD22222-2222-2222-2222-222222222222", total=300.0)
    pago = _complemento(
        "PAGO1111-1111-1111-1111-111111111111",
        [
            ("PPD11111-1111-1111-1111-111111111111", 500.0, 1),
            ("PPD22222-2222-2222-2222-222222222222", 300.0, 1),
        ],
    )
    db.agregar([p1, p2, pago])

    with db.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM pagos_relaciones WHERE cfdi_pago_uuid = ?",
            ("PAGO1111-1111-1111-1111-111111111111",),
        )
        assert cur.fetchone()[0] == 2


# ---------------------------------------------------------------------------
# Status (4 casos)
# ---------------------------------------------------------------------------


def test_status_sin_complemento(db):
    db.agregar([_ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0)])
    items = facturas_ppd(db)["items"]
    assert items[0]["status"] == "sin_complemento"
    assert items[0]["total_pagado"] == 0.0
    assert items[0]["saldo_pendiente"] == 1000.0


def test_status_pago_parcial(db):
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 400.0, 1)],
        ),
    ])
    items = facturas_ppd(db)["items"]
    ppd = next(i for i in items if i["uuid"] == "PPD11111-1111-1111-1111-111111111111")
    assert ppd["status"] == "pago_parcial"
    assert ppd["total_pagado"] == 400.0
    assert ppd["saldo_pendiente"] == 600.0


def test_status_pagado_completo_dos_parcialidades(db):
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 400.0, 1)],
        ),
        _complemento(
            "PAGO2222-2222-2222-2222-222222222222",
            [("PPD11111-1111-1111-1111-111111111111", 600.0, 2)],
        ),
    ])
    items = facturas_ppd(db)["items"]
    ppd = next(i for i in items if i["uuid"] == "PPD11111-1111-1111-1111-111111111111")
    assert ppd["status"] == "pagado_completo"
    assert ppd["num_pagos"] == 2


def test_status_sobrante(db):
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1100.0, 1)],
        ),
    ])
    items = facturas_ppd(db)["items"]
    ppd = next(i for i in items if i["uuid"] == "PPD11111-1111-1111-1111-111111111111")
    assert ppd["status"] == "sobrante"
    assert any("Sobrante" in w for w in ppd["warnings"])


# ---------------------------------------------------------------------------
# Huérfanos
# ---------------------------------------------------------------------------


def test_pagos_huerfanos(db):
    """Complemento referencia un UUID PPD que no está cargado."""
    db.agregar([
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("FANTASMA-NOT-LOADED-IN-DB-XXXXXXXX", 500.0, 1)],
        ),
    ])
    items = pagos_huerfanos(db)
    assert len(items) == 1
    assert items[0]["cfdi_pago_uuid"] == "PAGO1111-1111-1111-1111-111111111111"


# ---------------------------------------------------------------------------
# Extemporáneo
# ---------------------------------------------------------------------------


def test_complemento_extemporaneo(db):
    """FechaPago=15 enero, emisión=10 febrero → extemporáneo (límite: 5 feb)."""
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111"),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1160.0, 1)],
            fecha_pago="2026-01-15T10:00:00",
            fecha_emision="2026-02-10T10:00:00",  # 5 días tarde
        ),
    ])
    items = analisis_fechas(db)
    assert len(items) == 1
    assert items[0]["cfdi_pago_uuid"] == "PAGO1111-1111-1111-1111-111111111111"
    assert items[0]["dias_retraso"] >= 4


def test_complemento_a_tiempo_no_es_extemporaneo(db):
    """FechaPago=15 enero, emisión=3 febrero → a tiempo (antes del 5 feb)."""
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111"),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1160.0, 1)],
            fecha_pago="2026-01-15T10:00:00",
            fecha_emision="2026-02-03T10:00:00",
        ),
    ])
    assert analisis_fechas(db) == []


# ---------------------------------------------------------------------------
# Incidencias PUE+complemento
# ---------------------------------------------------------------------------


def test_incidencia_pue(db):
    """Factura emitida PUE referenciada por un complemento → riesgo fiscal."""
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", metodo_pago="PUE"),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1160.0, 1)],
        ),
    ])
    items = incidencias_pue(db)
    assert len(items) == 1
    assert items[0]["factura_uuid"] == "PPD11111-1111-1111-1111-111111111111"
    assert "duplica" in items[0]["descripcion_riesgo"]


# ---------------------------------------------------------------------------
# Stats globales
# ---------------------------------------------------------------------------


def test_stats_pagos_kpis(db):
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0),
        _ppd("PPD22222-2222-2222-2222-222222222222", total=2000.0),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1000.0, 1)],
        ),
    ])
    s = stats_pagos(db)
    assert s["total_ingresos_ppd"] == 2
    assert s["pagos_completos"] == 1
    assert s["sin_complemento"] == 1
    assert s["total_pagos"] == 1
    assert s["porcentaje_conciliados"] == 50.0
    assert s["monto_total_sin_pagar"] == 2000.0
    assert s["total_global_ppd"] == 2


# ---------------------------------------------------------------------------
# Drilldown
# ---------------------------------------------------------------------------


def test_detalle_pagos_de_ppd_ordena_por_parcialidad(db):
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0),
        _complemento(
            "PAGO2222-2222-2222-2222-222222222222",
            [("PPD11111-1111-1111-1111-111111111111", 600.0, 2)],
        ),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 400.0, 1)],
        ),
    ])
    detalle = detalle_pagos_de_ppd(db, "PPD11111-1111-1111-1111-111111111111")
    assert [p["docto_num_parcialidad"] for p in detalle] == [1, 2]


# ---------------------------------------------------------------------------
# Export XLSX
# ---------------------------------------------------------------------------


def test_to_xlsx_genera_sheets_basicos(db):
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", total=1000.0),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1000.0, 1)],
        ),
    ])
    xlsx_bytes = to_xlsx(db)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    assert "Resumen" in wb.sheetnames
    assert "Facturas PPD" in wb.sheetnames
    assert "Detalle de pagos" in wb.sheetnames


def test_to_xlsx_incluye_alertas_condicionales(db):
    """Cuando hay incidencias PUE y huérfanos, los sheets condicionales aparecen."""
    db.agregar([
        _ppd("PPD11111-1111-1111-1111-111111111111", metodo_pago="PUE"),
        _complemento(
            "PAGO1111-1111-1111-1111-111111111111",
            [("PPD11111-1111-1111-1111-111111111111", 1160.0, 1)],
        ),
        _complemento(
            "PAGO2222-2222-2222-2222-222222222222",
            [("FANTASMA-NOT-LOADED-IN-DB-XXXXXXXX", 500.0, 1)],
        ),
    ])
    xlsx_bytes = to_xlsx(db)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    assert "Incidencias PUE" in wb.sheetnames
    assert "Pagos huérfanos" in wb.sheetnames


# ---------------------------------------------------------------------------
# Filtros persistidos con key separada
# ---------------------------------------------------------------------------


def test_filtros_pagos_separados_de_cfdi(db):
    """`pagos_actuales` y `actuales` son keys independientes."""
    db.filtros_set({"tipo": "I"}, key="actuales")  # filtros CFDI
    db.filtros_set({"status": ["pago_parcial"]}, key="pagos_actuales")  # filtros Pagos

    assert db.filtros_get(key="actuales") == {"tipo": "I"}
    assert db.filtros_get(key="pagos_actuales") == {"status": ["pago_parcial"]}
